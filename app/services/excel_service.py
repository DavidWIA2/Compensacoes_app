import glob
import os
import re
import shutil
import tempfile
import time
import uuid
from datetime import datetime
from typing import Callable, Dict, List, Optional, cast

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from app.models.compensacao import Compensacao
from app.models.display_columns import DISPLAY_COLUMNS
from app.models.plantio_item import PlantioItem
from app.services.plantio_service import (
    clone_plantios,
    legacy_plantios_from_record,
    normalize_plantios,
    record_plantio_items,
    sync_legacy_plantio_fields,
)
from app.services.records_service import storage_tipo_value
from app.utils.logger import get_logger


logger = get_logger("Excel")

MAX_BACKUPS = 10
BACKUP_FOLDER_NAME = "backups_historico"
SHEET_NAME = "Compensações"
PLANTIOS_SHEET_NAME = "Plantios"

EXPECTED_HEADERS = {
    **{attr: label for label, attr in DISPLAY_COLUMNS},
    "latitude_plantio": "Lat_Plantio",
    "longitude_plantio": "Lon_Plantio",
    "latitude": "Latitude",
    "longitude": "Longitude",
    "uid": "UID",
}
EXPECTED_HEADER_ALIASES = {
    "eletronico": ("Tipo", "Eletrônico"),
}

EXPECTED_PLANTIO_HEADERS = {
    "uid_registro": "UID_Registro",
    "sequence": "Sequencia",
    "endereco": "Endereco_Plantio",
    "qtd_mudas": "Qtd_Mudas",
    "latitude": "Lat_Plantio",
    "longitude": "Lon_Plantio",
}


class InvalidFileError(Exception):
    """Excecao levantada quando o arquivo nao e um Excel valido."""


class WorkbookModifiedExternallyError(RuntimeError):
    """Excecao levantada quando a planilha muda no disco apos ter sido carregada."""


class ExcelService:
    def __init__(self):
        self.path: Optional[str] = None
        self.wb: Optional[openpyxl.Workbook] = None
        self.ws: Optional[Worksheet] = None
        self.plantio_ws: Optional[Worksheet] = None
        self.col_map: Dict[str, int] = {}
        self.plantio_col_map: Dict[str, int] = {}
        self.uid_to_row: Dict[str, int] = {}
        self.last_backup_time = 0
        self.merged_cells_warning = False
        self.loaded_source_mtime_ns = 0
        self.loaded_source_size = 0

    @staticmethod
    def _read_file_identity(path: str) -> tuple[int, int]:
        normalized_path = os.path.abspath(str(path or "").strip())
        if not normalized_path or not os.path.exists(normalized_path):
            return 0, 0
        try:
            stat_result = os.stat(normalized_path)
        except OSError:
            return 0, 0
        return int(getattr(stat_result, "st_mtime_ns", 0) or 0), int(getattr(stat_result, "st_size", 0) or 0)

    def _refresh_loaded_file_identity(self) -> None:
        if not self.path:
            self.loaded_source_mtime_ns = 0
            self.loaded_source_size = 0
            return
        self.loaded_source_mtime_ns, self.loaded_source_size = self._read_file_identity(self.path)

    def ensure_workbook_is_current(self) -> None:
        if not self.path:
            return
        current_mtime_ns, current_size = self._read_file_identity(self.path)
        if current_mtime_ns <= 0 and current_size <= 0:
            raise FileNotFoundError(f"Arquivo nao encontrado: {self.path}")
        if self.loaded_source_mtime_ns <= 0 and self.loaded_source_size <= 0:
            self.loaded_source_mtime_ns = current_mtime_ns
            self.loaded_source_size = current_size
            return
        if current_mtime_ns != self.loaded_source_mtime_ns or current_size != self.loaded_source_size:
            raise WorkbookModifiedExternallyError(
                "A planilha foi alterada por outro processo desde a ultima carga. Recarregue antes de continuar."
            )

    def _is_valid_zip_signature(self, path: str) -> bool:
        try:
            with open(path, "rb") as handle:
                header = handle.read(4)
                return header == b"PK\x03\x04"
        except Exception:
            return False

    def _save_workbook(self) -> None:
        if not self.wb or not self.path:
            raise ValueError("Nenhuma planilha carregada para salvar.")

        target_dir = os.path.dirname(os.path.abspath(self.path)) or "."
        fd, tmp_path = tempfile.mkstemp(prefix="compensacoes_", suffix=".xlsx", dir=target_dir)
        os.close(fd)
        try:
            self.wb.save(tmp_path)
            os.replace(tmp_path, self.path)
            self._refresh_loaded_file_identity()
        except Exception:
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except OSError:
                logger.warning(f"[EXCEL] Nao foi possivel remover arquivo temporario: {tmp_path}")
            raise

    def load(self, path: str) -> List[Compensacao]:
        if not os.path.exists(path):
            raise FileNotFoundError(f"Arquivo nao encontrado: {path}")

        if not self._is_valid_zip_signature(path):
            raise InvalidFileError(
                f"O arquivo nao parece ser uma planilha Excel valida (.xlsx corrompido ou formato incorreto): {path}"
            )

        self.path = path
        from zipfile import BadZipFile

        try:
            self.wb = openpyxl.load_workbook(path, data_only=False)
        except BadZipFile:
            raise InvalidFileError("O arquivo Excel esta corrompido ou nao pode ser lido (Erro interno de ZIP).")
        except Exception as exc:
            raise RuntimeError(f"Erro inesperado ao abrir a planilha: {exc}")

        if SHEET_NAME in self.wb.sheetnames:
            self.ws = self.wb[SHEET_NAME]
        else:
            self.ws = self.wb.active

        self._ensure_tracking_headers()
        self._build_column_map()

        self.plantio_ws = self.wb[PLANTIOS_SHEET_NAME] if PLANTIOS_SHEET_NAME in self.wb.sheetnames else None
        if self.plantio_ws is not None:
            self._ensure_plantio_headers()
            self._build_plantio_column_map()
        else:
            self.plantio_col_map.clear()
        plantios_by_uid = self._load_plantios_by_uid()

        records: List[Compensacao] = []
        needs_save = False
        self.uid_to_row.clear()
        self.merged_cells_warning = False
        seen_uids = set()

        for row_idx, row_cells in enumerate(self.ws.iter_rows(min_row=2, values_only=False), start=2):
            def get_val(key: str):
                col = self.col_map.get(key)
                if col and col <= len(row_cells):
                    return row_cells[col - 1].value
                return None

            main_vals = [get_val(key) for key in ["oficio_processo", "eletronico", "caixa", "av_tec"]]
            if all(value is None or str(value).strip() == "" for value in main_vals):
                continue

            uid_val = self._str(get_val("uid"))
            if not uid_val or uid_val in seen_uids:
                uid_val = uuid.uuid4().hex
                col_uid = self.col_map.get("uid")
                if col_uid:
                    self.ws.cell(row=row_idx, column=col_uid).value = uid_val
                    needs_save = True

            seen_uids.add(uid_val)
            self.uid_to_row[uid_val] = row_idx

            record = Compensacao(
                excel_row=row_idx,
                oficio_processo=self._str(get_val("oficio_processo")),
                eletronico=storage_tipo_value(self._str(get_val("eletronico"))),
                caixa=self._str(get_val("caixa")),
                av_tec=self._str(get_val("av_tec")),
                compensacao=get_val("compensacao"),
                endereco=self._str(get_val("endereco")),
                microbacia=self._str(get_val("microbacia")),
                compensado=self._str(get_val("compensado")),
                endereco_plantio=self._str(get_val("endereco_plantio")),
                latitude_plantio=self._str(get_val("latitude_plantio")),
                longitude_plantio=self._str(get_val("longitude_plantio")),
                latitude=self._str(get_val("latitude")),
                longitude=self._str(get_val("longitude")),
                uid=uid_val,
            )
            record.plantios = clone_plantios(plantios_by_uid.get(uid_val) or legacy_plantios_from_record(record))
            sync_legacy_plantio_fields(record)
            records.append(record)

        if needs_save:
            try:
                self._save_workbook()
            except Exception as exc:
                logger.warning(f"[EXCEL] Nao foi possivel salvar novos UIDs gerados na leitura: {exc}")

        self._refresh_loaded_file_identity()

        return records

    def _build_column_map(self):
        self.col_map.clear()
        if not self.ws:
            return

        headers = [str(cell.value).strip() if cell.value else "" for cell in self.ws[1]]

        from app.services.records_service import remove_accents

        def normalize(text: str) -> str:
            return remove_accents(str(text)).strip().upper()

        normalized_headers = [normalize(header) for header in headers]
        for key, expected_name in EXPECTED_HEADERS.items():
            expected_names = EXPECTED_HEADER_ALIASES.get(key, (expected_name,))
            normalized_expected_names = [normalize(name) for name in expected_names]
            direct_match = next(
                (
                    normalized_headers.index(expected_name_candidate) + 1
                    for expected_name_candidate in normalized_expected_names
                    if expected_name_candidate in normalized_headers
                ),
                None,
            )
            if direct_match is not None:
                self.col_map[key] = direct_match
            else:
                found = False
                for index, header in enumerate(normalized_headers):
                    for normalized_expected_name in normalized_expected_names:
                        if normalized_expected_name in header or header in normalized_expected_name:
                            self.col_map[key] = index + 1
                            found = True
                            break
                    if found:
                        break
                if not found:
                    logger.warning(f"[EXCEL] Coluna '{expected_name}' nao mapeada.")

    def _build_plantio_column_map(self):
        self.plantio_col_map.clear()
        if not self.plantio_ws:
            return

        headers = [str(cell.value).strip() if cell.value else "" for cell in self.plantio_ws[1]]

        from app.services.records_service import remove_accents

        def normalize(text: str) -> str:
            return remove_accents(str(text)).strip().upper()

        normalized_headers = [normalize(header) for header in headers]
        for key, expected_name in EXPECTED_PLANTIO_HEADERS.items():
            norm_expected = normalize(expected_name)
            try:
                self.plantio_col_map[key] = normalized_headers.index(norm_expected) + 1
            except ValueError:
                logger.warning(f"[EXCEL] Coluna de plantio '{expected_name}' nao mapeada.")

    def _create_rotating_backup(self, *, force: bool = False, label: str = "") -> Optional[str]:
        if not self.path:
            return None

        now = time.time()
        if not force and now - self.last_backup_time < 300:
            return None
        self.last_backup_time = now

        base_dir = os.path.dirname(self.path)
        backup_dir = os.path.join(base_dir, BACKUP_FOLDER_NAME)
        os.makedirs(backup_dir, exist_ok=True)
        filename = os.path.basename(self.path)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S_%f")
        safe_label = re.sub(r"[^A-Za-z0-9._-]+", "-", str(label or "").strip()).strip("-._")
        suffix = f"_{safe_label}" if safe_label else ""
        backup_name = f"{os.path.splitext(filename)[0]}_{timestamp}{suffix}.xlsx"
        backup_path = os.path.join(backup_dir, backup_name)
        try:
            shutil.copy2(self.path, backup_path)
        except Exception as exc:
            logger.warning(f"[EXCEL] Nao foi possivel criar backup: {exc}")
            return None

        files = glob.glob(os.path.join(backup_dir, "*.xlsx"))
        files.sort(key=os.path.getmtime)
        while len(files) > MAX_BACKUPS:
            oldest = files.pop(0)
            try:
                os.remove(oldest)
            except Exception as exc:
                logger.warning(f"[EXCEL] Nao foi possivel remover backup antigo '{oldest}': {exc}")
        return backup_path

    def create_operation_backup(self, label: str) -> Optional[str]:
        return self._create_rotating_backup(force=True, label=label)

    def _find_row_by_uid(self, uid: str) -> Optional[int]:
        if not self.ws or not uid:
            return None

        col_uid = self.col_map.get("uid")
        if not col_uid:
            return None

        if uid in self.uid_to_row:
            row = self.uid_to_row[uid]
            try:
                if self.ws.cell(row=row, column=col_uid).value == uid:
                    return row
            except Exception:
                pass

        for row_idx, row_cells in enumerate(
            self.ws.iter_rows(min_row=2, min_col=col_uid, max_col=col_uid, values_only=True),
            start=2,
        ):
            if row_cells[0] == uid:
                self.uid_to_row[uid] = row_idx
                return row_idx
        return None

    def find_row_by_uid(self, uid: str) -> Optional[int]:
        return self._find_row_by_uid(uid)

    def _row_has_values(self, row_idx: int) -> bool:
        if not self.ws:
            return False

        for col_idx in range(1, self.ws.max_column + 1):
            value = self.ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            return True
        return False

    def add_new(self, c: Compensacao) -> int:
        self.ensure_workbook_is_current()
        self._create_rotating_backup()
        new_row = self._append_new_without_save(c)
        self._save_workbook()
        return new_row

    def save_edit(self, c: Compensacao):
        def mutate(shadow: "ExcelService") -> None:
            target_row = shadow._resolve_target_row(c, require_uid_match=True)
            shadow._write_row(target_row, c)
            shadow._sync_plantio_rows(c)

        self._commit_shadow_mutation(mutate)

    def save_batch_edits(self, records: List[Compensacao]) -> int:
        if not records:
            return 0

        def mutate(shadow: "ExcelService") -> int:
            updated = 0
            for record in records:
                target_row = shadow._resolve_target_row(record, require_uid_match=True)
                shadow._write_row(target_row, record)
                shadow._sync_plantio_rows(record)
                updated += 1
            return updated

        return cast(int, self._commit_shadow_mutation(mutate))

    def import_records_atomic(
        self,
        records: List[Compensacao],
        *,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> int:
        if not records:
            return 0

        def mutate(shadow: "ExcelService") -> int:
            total = len(records)
            imported = 0
            for record in records:
                shadow._append_new_without_save(record)
                imported += 1
                if progress_callback:
                    progress_callback(imported, total)
            return imported

        return cast(int, self._commit_shadow_mutation(mutate))

    def delete_record_shift_up(self, row_idx: int, uid: str = ""):
        self.ensure_workbook_is_current()
        self._create_rotating_backup()
        ws = self.ws
        if ws is None:
            raise ValueError("Nenhuma planilha carregada para excluir registros.")
        target_row = row_idx
        target_uid = uid
        if uid:
            found_row = self._find_row_by_uid(uid)
            if not found_row:
                raise LookupError("Nao foi possivel localizar o registro pelo UID. Recarregue a planilha e tente novamente.")
            target_row = found_row
            if uid in self.uid_to_row:
                del self.uid_to_row[uid]
        else:
            col_uid = self.col_map.get("uid")
            if col_uid:
                target_uid = self._str(ws.cell(row=target_row, column=col_uid).value)

        self._delete_plantio_rows_for_uid(target_uid)
        ws.delete_rows(target_row, 1)
        self.uid_to_row.clear()
        self._save_workbook()

    def read_all(self) -> List[Compensacao]:
        if not self.path:
            raise ValueError("Nenhum arquivo Excel carregado.")
        return self.load(self.path)

    def _write_row(self, row: int, c: Compensacao):
        ws = self.ws
        if ws is None:
            raise ValueError("Nenhuma planilha carregada para escrita.")
        sync_legacy_plantio_fields(c)
        data_map = {
            "oficio_processo": c.oficio_processo,
            "eletronico": storage_tipo_value(c.eletronico),
            "caixa": c.caixa,
            "av_tec": c.av_tec,
            "compensacao": c.compensacao,
            "endereco": c.endereco,
            "microbacia": c.microbacia,
            "compensado": c.compensado,
            "endereco_plantio": c.endereco_plantio,
            "latitude_plantio": c.latitude_plantio,
            "longitude_plantio": c.longitude_plantio,
            "latitude": c.latitude,
            "longitude": c.longitude,
            "uid": c.uid,
        }

        for key, value in data_map.items():
            col_idx = self.col_map.get(key)
            if not col_idx:
                continue

            cell = ws.cell(row=row, column=col_idx)
            if not isinstance(cell, MergedCell):
                cell.value = value
            else:
                self.merged_cells_warning = True

    def _append_new_without_save(self, c: Compensacao) -> int:
        new_row = 2
        while self._row_has_values(new_row):
            new_row += 1

        if not c.uid:
            c.uid = uuid.uuid4().hex
        c.excel_row = new_row
        self._write_row(new_row, c)
        self._sync_plantio_rows(c)
        self.uid_to_row[c.uid] = new_row
        return new_row

    def _resolve_target_row(self, c: Compensacao, *, require_uid_match: bool = False) -> int:
        target_row = c.excel_row
        if c.uid:
            found_row = self._find_row_by_uid(c.uid)
            if found_row:
                target_row = found_row
                c.excel_row = found_row
            elif require_uid_match:
                raise LookupError("Nao foi possivel localizar o registro pelo UID. Recarregue a planilha e tente novamente.")
        elif require_uid_match:
            raise LookupError("Nao foi possivel localizar o registro pelo UID. Recarregue a planilha e tente novamente.")
        return target_row

    def _adopt_runtime_state(self, other: "ExcelService") -> None:
        self.path = other.path
        self.wb = other.wb
        self.ws = other.ws
        self.plantio_ws = other.plantio_ws
        self.col_map = dict(other.col_map)
        self.plantio_col_map = dict(other.plantio_col_map)
        self.uid_to_row = dict(other.uid_to_row)
        self.last_backup_time = other.last_backup_time
        self.merged_cells_warning = other.merged_cells_warning
        self.loaded_source_mtime_ns = other.loaded_source_mtime_ns
        self.loaded_source_size = other.loaded_source_size

    def _build_shadow_service(self) -> "ExcelService":
        if not self.path:
            raise ValueError("Nenhuma planilha carregada para salvar.")

        self.ensure_workbook_is_current()
        shadow = ExcelService()
        shadow.load(self.path)
        shadow.last_backup_time = self.last_backup_time
        return shadow

    def _commit_shadow_mutation(self, mutation: Callable[["ExcelService"], object]) -> object:
        shadow = self._build_shadow_service()
        shadow._create_rotating_backup()
        result = mutation(shadow)
        shadow._save_workbook()
        self._adopt_runtime_state(shadow)
        return result

    def _ensure_tracking_headers(self):
        if not self.ws:
            return

        headers = [str(cell.value).strip() if cell.value else "" for cell in self.ws[1]]

        from app.services.records_service import remove_accents

        def normalize(text: str) -> str:
            return remove_accents(str(text)).strip().upper()

        normalized_headers = {normalize(header) for header in headers if header}
        for key, label in EXPECTED_HEADERS.items():
            alias_names = EXPECTED_HEADER_ALIASES.get(key, (label,))
            if not any(normalize(alias_name) in normalized_headers for alias_name in alias_names):
                new_col = len(headers) + 1
                self.ws.cell(row=1, column=new_col).value = label
                headers.append(label)
                normalized_headers.add(normalize(label))

    def _ensure_plantio_headers(self):
        if not self.plantio_ws:
            return

        headers = [str(cell.value).strip() if cell.value else "" for cell in self.plantio_ws[1]]
        if len(headers) == 1 and not headers[0]:
            headers = []

        from app.services.records_service import remove_accents

        def normalize(text: str) -> str:
            return remove_accents(str(text)).strip().upper()

        normalized_headers = {normalize(header) for header in headers if header}
        for key, label in EXPECTED_PLANTIO_HEADERS.items():
            if normalize(label) not in normalized_headers:
                new_col = len(headers) + 1
                self.plantio_ws.cell(row=1, column=new_col).value = label
                headers.append(label)
                normalized_headers.add(normalize(label))
        self.plantio_ws.sheet_state = "hidden"

    def _ensure_plantio_sheet(self):
        if not self.wb:
            return
        if self.plantio_ws is None:
            self.plantio_ws = (
                self.wb[PLANTIOS_SHEET_NAME]
                if PLANTIOS_SHEET_NAME in self.wb.sheetnames
                else self.wb.create_sheet(PLANTIOS_SHEET_NAME)
            )
        self._ensure_plantio_headers()
        self._build_plantio_column_map()

    def _load_plantios_by_uid(self) -> Dict[str, List[PlantioItem]]:
        plantios_by_uid: Dict[str, List[PlantioItem]] = {}
        if not self.plantio_ws or not self.plantio_col_map:
            return plantios_by_uid

        def get_val(row_cells, key: str):
            col = self.plantio_col_map.get(key)
            if col and col <= len(row_cells):
                return row_cells[col - 1].value
            return None

        for row_cells in self.plantio_ws.iter_rows(min_row=2, values_only=False):
            uid = self._str(get_val(row_cells, "uid_registro"))
            endereco = self._str(get_val(row_cells, "endereco"))
            qtd_mudas = self._str(get_val(row_cells, "qtd_mudas"))
            latitude = self._str(get_val(row_cells, "latitude"))
            longitude = self._str(get_val(row_cells, "longitude"))
            if not uid or not any([endereco, qtd_mudas, latitude, longitude]):
                continue

            sequence_raw = get_val(row_cells, "sequence")
            try:
                sequence = int(sequence_raw)
            except (TypeError, ValueError):
                sequence = len(plantios_by_uid.get(uid, [])) + 1

            plantios_by_uid.setdefault(uid, []).append(
                PlantioItem(
                    sequence=sequence,
                    endereco=endereco,
                    qtd_mudas=qtd_mudas,
                    latitude=latitude,
                    longitude=longitude,
                )
            )

        for uid, items in list(plantios_by_uid.items()):
            plantios_by_uid[uid] = normalize_plantios(sorted(items, key=lambda item: item.sequence))
        return plantios_by_uid

    def _delete_plantio_rows_for_uid(self, uid: str):
        if not uid or not self.plantio_ws or not self.plantio_col_map:
            return

        col_uid = self.plantio_col_map.get("uid_registro")
        if not col_uid:
            return

        rows_to_delete = []
        for row_idx, row_cells in enumerate(
            self.plantio_ws.iter_rows(min_row=2, min_col=col_uid, max_col=col_uid, values_only=True),
            start=2,
        ):
            if self._str(row_cells[0]) == uid:
                rows_to_delete.append(row_idx)

        for row_idx in reversed(rows_to_delete):
            self.plantio_ws.delete_rows(row_idx, 1)

    def _sync_plantio_rows(self, record: Compensacao):
        if not self.wb:
            return

        normalized = record_plantio_items(record)
        if not normalized and self.plantio_ws is None:
            return

        self._ensure_plantio_sheet()
        if not self.plantio_ws:
            return

        self._delete_plantio_rows_for_uid(record.uid)
        if not normalized:
            return

        for item in normalized:
            row_idx = self.plantio_ws.max_row + 1
            data_map = {
                "uid_registro": record.uid,
                "sequence": item.sequence,
                "endereco": item.endereco,
                "qtd_mudas": item.qtd_mudas,
                "latitude": item.latitude,
                "longitude": item.longitude,
            }
            for key, value in data_map.items():
                col_idx = self.plantio_col_map.get(key)
                if col_idx:
                    self.plantio_ws.cell(row=row_idx, column=col_idx).value = value

    def _str(self, value) -> str:
        return "" if value is None else str(value).strip()
