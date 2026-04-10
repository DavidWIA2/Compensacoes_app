from __future__ import annotations

import os
from datetime import datetime
from functools import partial
from typing import Any, List, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.platypus import HRFlowable, Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models.compensacao import Compensacao
from app.models.display_columns import DISPLAY_COLUMNS
from app.services.ficha_report_service import export_individual_pdf as export_ficha_report_pdf
from app.services.report_service_support import (
    DATA_SHEET_NAME,
    INSTITUTIONAL_APP_NAME,
    INSTITUTIONAL_REPORT_SUBTITLE,
    INSTITUTIONAL_SOURCE_LABEL,
    REPORT_DETAIL_LABEL,
    REPORT_SUMMARY_LABEL,
    REPORT_TITLE,
    SUMMARY_LABEL_FILTERS,
    SUMMARY_LABEL_GENERAL,
    SUMMARY_LABEL_METRIC,
    SUMMARY_LABEL_PENDING_MICRO,
    SUMMARY_LABEL_PENDING_TYPE,
    SUMMARY_LABEL_VALUE,
    SUMMARY_SHEET_NAME,
    ReportMetadataRow,
    build_dashboard_chart_rows,
    build_department_header_html_lines,
    build_grid_pdf_layout,
    build_individual_report_rows,
    build_records_to_dict_list,
    build_report_metadata_rows,
    build_selected_headers,
    format_individual_status,
    resolve_report_logo_path,
)


ALL_COLUMNS = DISPLAY_COLUMNS

BRAND_BLUE = "1F4E79"
BRAND_BLUE_LIGHT = "EAF2FA"
BRAND_BORDER = "C8D5E3"
BRAND_FILL = PatternFill(start_color=BRAND_BLUE_LIGHT, end_color=BRAND_BLUE_LIGHT, fill_type="solid")
HEADER_FILL = PatternFill(start_color=BRAND_BLUE, end_color=BRAND_BLUE, fill_type="solid")
NEUTRAL_FILL = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="EEF6ED", end_color="EEF6ED", fill_type="solid")
BORDER_STYLE = Border(
    left=Side(style="thin", color=BRAND_BORDER),
    right=Side(style="thin", color=BRAND_BORDER),
    top=Side(style="thin", color=BRAND_BORDER),
    bottom=Side(style="thin", color=BRAND_BORDER),
)
FONT_BOLD = Font(bold=True)
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=15, color=BRAND_BLUE)
SUBTITLE_FONT = Font(italic=True, color="5B6B7B")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _resolve_report_logo_path() -> str:
    return resolve_report_logo_path()


def _selected_headers(selected_cols: List[str]) -> List[str]:
    return build_selected_headers(selected_cols)


def _format_individual_status(record: Compensacao) -> str:
    return format_individual_status(record)


def _build_individual_pdf_rows(record: Compensacao, observation: str = "") -> List[List[str]]:
    return build_individual_report_rows(record, observation)


def _records_to_dict_list(records: List[Compensacao], selected_cols: List[str]) -> List[dict[str, str]]:
    return build_records_to_dict_list(records, selected_cols)


def _style_worksheet(ws, highlight_compensado: bool = False):
    col_compensado_idx = -1
    if highlight_compensado:
        for col_idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).strip().lower() == "compensado":
                col_compensado_idx = col_idx
                break

    for row in ws.iter_rows():
        is_green_row = False
        if col_compensado_idx != -1 and row[0].row > 1:
            value = row[col_compensado_idx - 1].value
            if value and str(value).strip().upper() == "SIM":
                is_green_row = True

        for cell in row:
            cell.border = BORDER_STYLE
            if cell.row == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = ALIGN_CENTER
            elif is_green_row:
                cell.fill = HIGHLIGHT_FILL
                cell.alignment = ALIGN_LEFT_WRAP
            else:
                cell.alignment = ALIGN_LEFT_WRAP

            if cell.row > 1 and cell.value and len(str(cell.value)) < 15:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        final_width = min(max(length + 2, 10), 60)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = final_width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _write_summary_table(worksheet, *, start_row: int, start_col: int, title: str, headers: tuple[str, str], rows):
    worksheet.cell(row=start_row, column=start_col, value=title).font = FONT_BOLD
    worksheet.cell(row=start_row, column=start_col).fill = BRAND_FILL
    worksheet.cell(row=start_row + 1, column=start_col, value=headers[0]).font = FONT_BOLD
    worksheet.cell(row=start_row + 1, column=start_col + 1, value=headers[1]).font = FONT_BOLD
    worksheet.cell(row=start_row + 1, column=start_col).fill = NEUTRAL_FILL
    worksheet.cell(row=start_row + 1, column=start_col + 1).fill = NEUTRAL_FILL
    worksheet.cell(row=start_row + 1, column=start_col).border = BORDER_STYLE
    worksheet.cell(row=start_row + 1, column=start_col + 1).border = BORDER_STYLE
    for index, row in enumerate(rows):
        worksheet.cell(row=start_row + 2 + index, column=start_col, value=row[0])
        worksheet.cell(row=start_row + 2 + index, column=start_col + 1, value=row[1])
        worksheet.cell(row=start_row + 2 + index, column=start_col).border = BORDER_STYLE
        worksheet.cell(row=start_row + 2 + index, column=start_col + 1).border = BORDER_STYLE


def _apply_workbook_properties(workbook, *, title: str, subject: str, description: str) -> None:
    workbook.properties.creator = INSTITUTIONAL_APP_NAME
    workbook.properties.lastModifiedBy = INSTITUTIONAL_APP_NAME
    workbook.properties.title = title
    workbook.properties.subject = subject
    workbook.properties.description = description
    workbook.properties.keywords = "compensações, meio ambiente, prefeitura, são carlos"
    workbook.properties.category = "Relatório institucional"


def _style_summary_worksheet(worksheet, *, metadata_rows: tuple[ReportMetadataRow, ...]) -> None:
    worksheet.merge_cells("A1:H1")
    worksheet["A1"] = REPORT_TITLE
    worksheet["A1"].font = TITLE_FONT
    worksheet["A1"].alignment = ALIGN_CENTER

    worksheet.merge_cells("A2:H2")
    worksheet["A2"] = f"{INSTITUTIONAL_REPORT_SUBTITLE} | {INSTITUTIONAL_APP_NAME}"
    worksheet["A2"].font = SUBTITLE_FONT
    worksheet["A2"].alignment = ALIGN_CENTER

    start_row = 4
    for index, item in enumerate(metadata_rows):
        row = start_row + index
        worksheet.cell(row=row, column=1, value=item.label).font = FONT_BOLD
        worksheet.cell(row=row, column=1).fill = BRAND_FILL
        worksheet.cell(row=row, column=1).border = BORDER_STYLE
        worksheet.cell(row=row, column=2, value=item.value).border = BORDER_STYLE
        worksheet.cell(row=row, column=2).alignment = ALIGN_LEFT_WRAP

    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 56
    worksheet.column_dimensions["E"].width = 24
    worksheet.column_dimensions["F"].width = 14
    worksheet.column_dimensions["I"].width = 20
    worksheet.column_dimensions["J"].width = 14
    worksheet.freeze_panes = "A10"


def _build_pdf_header(styles, *, title: str, metadata_rows: tuple[ReportMetadataRow, ...]):
    logo_path = _resolve_report_logo_path()
    logo = Spacer(1, 1)
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=1.05 * inch, height=0.88 * inch)

    header_title = ParagraphStyle(
        "InstitutionTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=14,
        alignment=1,
        textColor=colors.HexColor("#143A5A"),
    )
    header_text = ParagraphStyle(
        "InstitutionText",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        alignment=1,
        textColor=colors.HexColor("#334E68"),
    )
    report_title = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=18,
        alignment=0,
        textColor=colors.HexColor("#143A5A"),
        spaceAfter=8,
    )
    metadata_label = ParagraphStyle(
        "ReportMetaLabel",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#143A5A"),
    )
    metadata_value = ParagraphStyle(
        "ReportMetaValue",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#334E68"),
    )

    department_lines = [
        Paragraph(build_department_header_html_lines()[0], header_title),
        *[Paragraph(line, header_text) for line in build_department_header_html_lines()[1:]],
    ]
    header_table = Table([[logo, department_lines]], colWidths=[1.25 * inch, 8.2 * inch])
    header_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (0, 0), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )

    metadata_table = Table(
        [[Paragraph(item.label, metadata_label), Paragraph(item.value, metadata_value)] for item in metadata_rows],
        colWidths=[1.15 * inch, 7.85 * inch],
    )
    metadata_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EAF2FA")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#C8D5E3")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C8D5E3")),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    return [
        header_table,
        Spacer(1, 0.1 * inch),
        HRFlowable(width="100%", thickness=0.8, color=colors.HexColor("#90A4B8")),
        Spacer(1, 0.12 * inch),
        Paragraph(title, report_title),
        metadata_table,
        Spacer(1, 0.18 * inch),
    ]


def _build_kpi_table(kpis: List[Tuple[str, Any]]):
    rows = [[SUMMARY_LABEL_METRIC, SUMMARY_LABEL_VALUE]] + [[str(key), str(value)] for key, value in kpis]
    table = Table(rows, colWidths=[3.2 * inch, 1.4 * inch])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#9FB3C8")),
                ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#C8D5E3")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F7F9FC")),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    return table


def _build_dashboard_kpi_table(kpi_lines: List[str], styles):
    item_style = ParagraphStyle(
        "DashboardKpiItem",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#243B53"),
    )
    title_style = ParagraphStyle(
        "DashboardKpiTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11,
        textColor=colors.white,
    )
    rows = [[Paragraph(REPORT_SUMMARY_LABEL, title_style)]]
    rows.extend([[Paragraph(str(line), item_style)] for line in kpi_lines])
    table = Table(rows, colWidths=[8.9 * inch])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#9FB3C8")),
                ("INNERGRID", (0, 1), (-1, -1), 0.4, colors.HexColor("#D7E2EC")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return table


def _draw_pdf_page_frame(canvas, doc, *, title: str, generated_label: str):
    canvas.saveState()
    canvas.setStrokeColor(colors.HexColor("#C8D5E3"))
    canvas.line(doc.leftMargin, 18, doc.pagesize[0] - doc.rightMargin, 18)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#486581"))
    canvas.drawString(doc.leftMargin, 8, f"{INSTITUTIONAL_APP_NAME} | {title}")
    right_text = f"Emitido em {generated_label} | Página {canvas.getPageNumber()}"
    text_width = stringWidth(right_text, "Helvetica", 8)
    canvas.drawString(doc.pagesize[0] - doc.rightMargin - text_width, 8, right_text)
    canvas.restoreState()


def export_csv(path: str, records: List[Compensacao], selected_cols: List[str]):
    data = _records_to_dict_list(records, selected_cols)
    if not data:
        pd.DataFrame(columns=_selected_headers(selected_cols)).to_csv(
            path,
            index=False,
            encoding="utf-8-sig",
            sep=";",
        )
        return

    pd.DataFrame(data).to_csv(path, index=False, encoding="utf-8-sig", sep=";")


def export_excel_two_sheets(
    path: str,
    records: List[Compensacao],
    filtros_txt: str,
    selected_cols: List[str],
    kpis: List[Tuple[str, Any]],
    pend_micro_sorted: List[Tuple[str, float]],
    pend_ele_sorted: List[Tuple[str, float]],
):
    data = _records_to_dict_list(records, selected_cols)
    df_dados = pd.DataFrame(data) if data else pd.DataFrame(columns=_selected_headers(selected_cols))
    metadata_rows = build_report_metadata_rows(filtros_txt, source_label=INSTITUTIONAL_SOURCE_LABEL)

    try:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df_dados.to_excel(writer, sheet_name=DATA_SHEET_NAME, index=False)
            worksheet = writer.book.create_sheet(SUMMARY_SHEET_NAME)
            writer.sheets[SUMMARY_SHEET_NAME] = worksheet

            _apply_workbook_properties(
                writer.book,
                title=REPORT_TITLE,
                subject="Resumo gerencial de compensações",
                description="Exportação institucional da Plataforma de Gestão Ambiental.",
            )

            _style_summary_worksheet(worksheet, metadata_rows=metadata_rows)

            _write_summary_table(
                worksheet,
                start_row=10,
                start_col=1,
                title=SUMMARY_LABEL_GENERAL,
                headers=(SUMMARY_LABEL_METRIC, SUMMARY_LABEL_VALUE),
                rows=kpis,
            )
            _write_summary_table(
                worksheet,
                start_row=10,
                start_col=5,
                title=SUMMARY_LABEL_PENDING_MICRO,
                headers=("Microbacia", "Mudas"),
                rows=pend_micro_sorted,
            )
            _write_summary_table(
                worksheet,
                start_row=10,
                start_col=9,
                title=SUMMARY_LABEL_PENDING_TYPE,
                headers=("Tipo", "Mudas"),
                rows=pend_ele_sorted,
            )

            _style_worksheet(writer.sheets[DATA_SHEET_NAME], highlight_compensado=True)
    except Exception as exc:
        raise RuntimeError(f"Erro ao salvar Excel: {exc}") from exc


export_spreadsheet_two_sheets = export_excel_two_sheets


def export_pdf(
    path: str,
    records: List[Compensacao],
    filtros_txt: str,
    selected_cols: List[str],
    kpis: List[Tuple[str, Any]],
    pend_micro_sorted: List[Tuple[str, float]],
):
    del pend_micro_sorted

    doc = SimpleDocTemplate(
        path,
        pagesize=landscape(A4),
        rightMargin=22,
        leftMargin=22,
        topMargin=22,
        bottomMargin=26,
    )
    elements = []
    styles = getSampleStyleSheet()
    generated_label = datetime.now().strftime("%d/%m/%Y %H:%M")
    metadata_rows = build_report_metadata_rows(filtros_txt, source_label=INSTITUTIONAL_SOURCE_LABEL)

    style_header = ParagraphStyle(
        "HeaderStyle",
        parent=styles["Normal"],
        fontSize=7,
        leading=8,
        textColor=colors.whitesmoke,
        fontName="Helvetica-Bold",
        alignment=1,
    )
    style_cell = ParagraphStyle(
        "CellStyle",
        parent=styles["Normal"],
        fontSize=7,
        leading=8,
        alignment=0,
        textColor=colors.HexColor("#243B53"),
    )
    style_cell_center = ParagraphStyle(
        "CellStyleCenter",
        parent=style_cell,
        alignment=1,
    )
    section_title = ParagraphStyle(
        "SectionTitle",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#143A5A"),
        spaceAfter=6,
    )

    elements.extend(_build_pdf_header(styles, title=REPORT_TITLE, metadata_rows=metadata_rows))
    elements.append(Paragraph(REPORT_SUMMARY_LABEL, section_title))
    elements.append(_build_kpi_table(kpis))
    elements.append(Spacer(1, 0.15 * inch))

    page_width = landscape(A4)[0] - doc.leftMargin - doc.rightMargin
    layout = build_grid_pdf_layout(selected_cols, page_width)

    table_data = [[Paragraph(header, style_header) for header in layout.headers]]
    raw_data = _records_to_dict_list(records, selected_cols)
    for row_dict in raw_data:
        row_elements = []
        for index, _attr in enumerate(selected_cols):
            value = str(row_dict.get(layout.headers[index], "") or "")
            current_style = style_cell_center if layout.weights[index] <= 1.0 else style_cell
            row_elements.append(Paragraph(value, current_style))
        table_data.append(row_elements)

    main_table = Table(table_data, colWidths=list(layout.column_widths), repeatRows=1)
    main_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#B8C7D9")),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D7E2EC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )

    elements.append(Paragraph(REPORT_DETAIL_LABEL, section_title))
    elements.append(main_table)
    doc.build(
        elements,
        onFirstPage=partial(_draw_pdf_page_frame, title=REPORT_TITLE, generated_label=generated_label),
        onLaterPages=partial(_draw_pdf_page_frame, title=REPORT_TITLE, generated_label=generated_label),
    )


def export_dashboard_pdf(path: str, titulo: str, kpi_lines: List[str], filtros_txt: str, chart_images: List[str]):
    doc = SimpleDocTemplate(
        path,
        pagesize=landscape(A4),
        rightMargin=22,
        leftMargin=22,
        topMargin=22,
        bottomMargin=26,
    )
    elements = []
    styles = getSampleStyleSheet()
    generated_label = datetime.now().strftime("%d/%m/%Y %H:%M")
    metadata_rows = build_report_metadata_rows(
        filtros_txt,
        source_label="Painel executivo",
    )
    elements.extend(_build_pdf_header(styles, title=titulo, metadata_rows=metadata_rows))
    elements.append(_build_dashboard_kpi_table(kpi_lines, styles))
    elements.append(Spacer(1, 0.18 * inch))

    chart_objects = []
    for img_path in chart_images:
        if os.path.exists(img_path):
            chart_objects.append(Image(img_path, width=350, height=250))

    chart_rows = build_dashboard_chart_rows(chart_objects)
    if chart_rows:
        charts_table = Table(list(chart_rows))
        charts_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#C8D5E3")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E1E8F0")),
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FBFCFE")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ]
            )
        )
        elements.append(charts_table)

    doc.build(
        elements,
        onFirstPage=partial(_draw_pdf_page_frame, title=titulo, generated_label=generated_label),
        onLaterPages=partial(_draw_pdf_page_frame, title=titulo, generated_label=generated_label),
    )


def export_individual_pdf(filepath: str, record: Compensacao, observation: str = ""):
    return export_ficha_report_pdf(filepath, record, observation)
