from __future__ import annotations

import ast
import math
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Sequence

import openpyxl

from app.models.tcra import Tcra
from app.models.tcra_evento import TcraEvento
from app.services.tcra_records_service import normalize_status_label
from app.services.tcra_sqlite_service import TcraSqliteService


TCRA_SHEET_NAME = "TCRAs"
TCRA_SHEET_ALIASES = (TCRA_SHEET_NAME, "TCRA's")


def _stringify(value: object) -> str:
    return str(value or "").strip()


def _normalize_header(value: object) -> str:
    text = _stringify(value)
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(char for char in normalized if not unicodedata.combining(char)).casefold()


class _SafeNumericExpression(ast.NodeVisitor):
    _ALLOWED_BINARY = {
        ast.Add: lambda left, right: left + right,
        ast.Sub: lambda left, right: left - right,
        ast.Mult: lambda left, right: left * right,
        ast.Div: lambda left, right: left / right,
    }
    _ALLOWED_UNARY = {
        ast.UAdd: lambda value: value,
        ast.USub: lambda value: -value,
    }

    def visit_Expression(self, node: ast.Expression) -> float:
        return self.visit(node.body)

    def visit_BinOp(self, node: ast.BinOp) -> float:
        operator = self._ALLOWED_BINARY.get(type(node.op))
        if operator is None:
            raise ValueError("Operador nao suportado na formula numerica.")
        return float(operator(self.visit(node.left), self.visit(node.right)))

    def visit_UnaryOp(self, node: ast.UnaryOp) -> float:
        operator = self._ALLOWED_UNARY.get(type(node.op))
        if operator is None:
            raise ValueError("Operador unario nao suportado na formula numerica.")
        return float(operator(self.visit(node.operand)))

    def visit_Constant(self, node: ast.Constant) -> float:
        if isinstance(node.value, (int, float)):
            return float(node.value)
        raise ValueError("Constante nao suportada na formula numerica.")

    def generic_visit(self, node: ast.AST) -> float:
        raise ValueError("Expressao nao suportada na formula numerica.")


@dataclass(frozen=True)
class TcraImportIssue:
    row_index: int
    severity: str
    code: str
    message: str


@dataclass(frozen=True)
class TcraWorkbookAnalysis:
    workbook_path: Path
    worksheet_name: str
    importable_count: int
    skipped_count: int
    missing_columns: tuple[str, ...] = ()
    issues: tuple[TcraImportIssue, ...] = ()
    tcras: tuple[Tcra, ...] = ()

    def severity_counts(self) -> tuple[tuple[str, int], ...]:
        counts: dict[str, int] = {}
        for issue in self.issues:
            severity = _stringify(issue.severity).lower() or "warning"
            counts[severity] = counts.get(severity, 0) + 1
        return tuple(sorted(counts.items(), key=lambda item: (item[0], item[1])))

    def issue_code_counts(self) -> tuple[tuple[str, int], ...]:
        counts: dict[str, int] = {}
        for issue in self.issues:
            code = _stringify(issue.code) or "sem_codigo"
            counts[code] = counts.get(code, 0) + 1
        return tuple(sorted(counts.items(), key=lambda item: (-item[1], item[0])))

    def import_labels(self, *, limit: int = 5) -> tuple[str, ...]:
        labels: list[str] = []
        for tcra in self.tcras[: max(int(limit or 0), 0)]:
            label = _stringify(tcra.numero_tcra or tcra.numero_processo or tcra.local or tcra.uid)
            if label:
                labels.append(label)
        return tuple(labels)

    def summary_lines(self, *, max_issue_codes: int = 4) -> tuple[str, ...]:
        lines = [
            f"TCRAs importaveis: {self.importable_count}",
            f"Linhas descartadas: {self.skipped_count}",
            f"Avisos encontrados: {len(self.issues)}",
        ]
        if self.missing_columns:
            lines.append("Colunas ausentes: " + ", ".join(self.missing_columns))
        issue_counts = self.issue_code_counts()
        if issue_counts:
            top_codes = " | ".join(f"{code}: {count}" for code, count in issue_counts[: max_issue_codes or 0])
            lines.append("Principais ocorrencias: " + top_codes)
        import_labels = self.import_labels(limit=4)
        if import_labels:
            lines.append("Primeiros termos: " + " | ".join(import_labels))
        return tuple(lines)


@dataclass(frozen=True)
class TcraImportMergeResult:
    importable_count: int
    created_count: int
    updated_count: int
    imported_uids: tuple[str, ...] = ()

    def summary_lines(self) -> tuple[str, ...]:
        return (
            f"Merge importado: {self.importable_count}",
            f"Novos termos: {self.created_count}",
            f"Atualizados no banco: {self.updated_count}",
        )


class TcraExcelService:
    HEADER_ALIASES = {
        "numero_processo": ("Processo", "No Processo", "Numero Processo"),
        "local": ("Local",),
        "endereco": ("Endereco",),
        "data_proximo_relatorio": ("Relat. Periodico", "Relatório Periódico", "Próx. Relatório"),
        "data_ultimo_relatorio": ("Último Rel.", "Último Relatório"),
        "prazo_final": ("Prazo", "Prazo Final"),
        "servicos_exigidos": ("Servicos a realizar", "Servicos"),
        "area_m2": ("Tamanho", "Area", "Area m2"),
        "numero_mudas_previsto": ("No de Mudas", "Numero de Mudas"),
        "acompanhamento": ("Acompanhamento",),
        "observacoes": ("Observacoes", ""),
        "mpsp_relacionado": ("MPSP?", "MPSP"),
    }

    def __init__(self, *, sqlite_service: TcraSqliteService | None = None, today: date | None = None):
        self.sqlite_service = sqlite_service or TcraSqliteService()
        self.today = today or date.today()
        self.header_aliases = dict(self.HEADER_ALIASES)

    def load_workbook(self, path: str | Path) -> list[Tcra]:
        return list(self.analyze_workbook(path).tcras)

    def analyze_workbook(self, path: str | Path) -> TcraWorkbookAnalysis:
        workbook_path = Path(path)
        if not workbook_path.exists():
            raise FileNotFoundError(f"Arquivo nao encontrado: {workbook_path}")

        workbook = openpyxl.load_workbook(workbook_path, data_only=False)
        worksheet_name = self._resolve_worksheet_name(tuple(workbook.sheetnames))
        if not worksheet_name:
            expected_names = "', '".join(TCRA_SHEET_ALIASES)
            raise KeyError(f"Nenhuma aba de TCRA valida ('{expected_names}') existe em {workbook_path}.")

        worksheet = workbook[worksheet_name]
        column_map = self._build_column_map(worksheet)
        missing_columns = tuple(
            field_name
            for field_name in (
                "numero_processo",
                "local",
                "endereco",
                "acompanhamento",
                "prazo_final",
                "data_proximo_relatorio",
            )
            if field_name not in column_map
        )

        tcras: list[Tcra] = []
        issues: list[TcraImportIssue] = []
        skipped_count = 0
        for row_index, row_values in enumerate(
            worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=False),
            start=2,
        ):
            row_data = {key: self._read_cell(row_values, column_map.get(key)) for key in column_map}
            issues.extend(self._inspect_row(row_index=row_index, row_data=row_data))
            tcra = self._row_to_tcra(row_index=row_index, row_data=row_data)
            if tcra is None:
                if self._row_has_meaningful_data(row_values):
                    skipped_count += 1
                    issues.append(
                        TcraImportIssue(
                            row_index=row_index,
                            severity="warning",
                            code="linha_descartada_sem_identificacao",
                            message=(
                                "Linha com dados residuais foi descartada por nao ter processo, local nem endereco."
                            ),
                        )
                    )
                continue
            tcras.append(tcra)

        return TcraWorkbookAnalysis(
            workbook_path=workbook_path,
            worksheet_name=worksheet_name,
            importable_count=len(tcras),
            skipped_count=skipped_count,
            missing_columns=missing_columns,
            issues=tuple(issues),
            tcras=tuple(tcras),
        )

    def import_workbook(self, path: str | Path) -> int:
        analysis = self.analyze_workbook(path)
        return self.sqlite_service.replace_all(list(analysis.tcras))

    def merge_workbook(self, source: str | Path | TcraWorkbookAnalysis) -> TcraImportMergeResult:
        analysis = source if isinstance(source, TcraWorkbookAnalysis) else self.analyze_workbook(source)
        created_count = 0
        updated_count = 0
        imported_uids: list[str] = []

        for imported_record in analysis.tcras:
            existing = self.sqlite_service.find_duplicate_tcra(
                numero_processo=imported_record.numero_processo,
                numero_tcra=imported_record.numero_tcra,
                local=imported_record.local,
            )
            if existing is None:
                saved_uid = self.sqlite_service.upsert_tcra(imported_record)
                created_count += 1
            else:
                saved_uid = self.sqlite_service.upsert_tcra(self._merge_records(existing, imported_record))
                updated_count += 1
            imported_uids.append(saved_uid)

        return TcraImportMergeResult(
            importable_count=analysis.importable_count,
            created_count=created_count,
            updated_count=updated_count,
            imported_uids=tuple(imported_uids),
        )

    @staticmethod
    def _resolve_worksheet_name(sheet_names: Sequence[str]) -> str:
        normalized_map = {_normalize_header(name): name for name in sheet_names}
        for candidate in TCRA_SHEET_ALIASES:
            match = normalized_map.get(_normalize_header(candidate))
            if match:
                return match
        return ""

    @staticmethod
    def _row_has_meaningful_data(row_values: Sequence[openpyxl.cell.cell.Cell]) -> bool:
        return any(_stringify(cell.value) for cell in row_values if cell.value is not None)

    def _inspect_row(self, *, row_index: int, row_data: dict[str, object]) -> list[TcraImportIssue]:
        issues: list[TcraImportIssue] = []
        numero_processo = _stringify(row_data.get("numero_processo"))
        local = _stringify(row_data.get("local"))
        endereco = _stringify(row_data.get("endereco"))
        acompanhamento = _stringify(row_data.get("acompanhamento"))
        numero_mudas = _stringify(row_data.get("numero_mudas_previsto"))
        data_ultimo_relatorio = self._parse_date(row_data.get("data_ultimo_relatorio"))
        data_proximo_relatorio = self._parse_date(row_data.get("data_proximo_relatorio"))

        if (local or endereco) and not numero_processo:
            issues.append(
                TcraImportIssue(
                    row_index=row_index,
                    severity="warning",
                    code="processo_ausente",
                    message="Linha importada sem numero de processo; o acompanhamento ficara dependente do local/endereco.",
                )
            )

        if acompanhamento and self._looks_like_coordinate(acompanhamento):
            issues.append(
                TcraImportIssue(
                    row_index=row_index,
                    severity="warning",
                    code="acompanhamento_deslocado",
                    message="A coluna de acompanhamento parece conter coordenada; a linha pode estar deslocada.",
                )
            )

        if numero_mudas and self._looks_like_coordinate(numero_mudas):
            issues.append(
                TcraImportIssue(
                    row_index=row_index,
                    severity="warning",
                    code="numero_mudas_invalido",
                    message="A coluna de numero de mudas parece conter coordenada e foi desconsiderada.",
                )
            )

        if (
            data_ultimo_relatorio is not None
            and data_proximo_relatorio is not None
            and data_proximo_relatorio < data_ultimo_relatorio
        ):
            issues.append(
                TcraImportIssue(
                    row_index=row_index,
                    severity="warning",
                    code="datas_relatorio_inconsistentes",
                    message="O proximo relatorio esta antes do ultimo relatorio informado.",
                )
            )

        return issues

    def _build_column_map(self, worksheet) -> dict[str, int]:
        headers = [worksheet.cell(row=1, column=index).value for index in range(1, worksheet.max_column + 1)]
        normalized_headers = {_normalize_header(header): index for index, header in enumerate(headers, start=1)}
        column_map: dict[str, int] = {}
        for field_name, aliases in self.header_aliases.items():
            for alias in aliases:
                normalized_alias = _normalize_header(alias)
                if normalized_alias in normalized_headers:
                    column_map[field_name] = normalized_headers[normalized_alias]
                    break
        return column_map

    @staticmethod
    def _read_cell(row_values: Sequence[openpyxl.cell.cell.Cell], column_index: int | None) -> object:
        if column_index is None or column_index <= 0 or column_index > len(row_values):
            return None
        return row_values[column_index - 1].value

    def _row_to_tcra(self, *, row_index: int, row_data: dict[str, object]) -> Tcra | None:
        numero_processo = _stringify(row_data.get("numero_processo"))
        local = _stringify(row_data.get("local"))
        endereco = _stringify(row_data.get("endereco"))
        if not any([numero_processo, local, endereco]):
            return None

        area_m2 = self._parse_area(row_data.get("area_m2"))
        observacoes = self._normalize_note_text(row_data.get("observacoes"))
        acompanhamento = _stringify(row_data.get("acompanhamento"))
        data_ultimo_relatorio = self._parse_date(row_data.get("data_ultimo_relatorio"))
        data_proximo_relatorio = self._parse_date(row_data.get("data_proximo_relatorio"))
        prazo_final = self._parse_date(row_data.get("prazo_final"))
        servicos_exigidos = _stringify(row_data.get("servicos_exigidos"))
        status = self._infer_status(
            acompanhamento=acompanhamento,
            observacoes=observacoes,
            servicos=servicos_exigidos,
            prazo_final=prazo_final,
            data_proximo_relatorio=data_proximo_relatorio,
        )

        return Tcra(
            uid=f"tcra-import-{row_index}-{numero_processo or local or 'sem-id'}".lower().replace(" ", "-"),
            numero_processo=numero_processo,
            numero_tcra="",
            local=local,
            endereco=endereco,
            bairro=self._extract_bairro(endereco),
            orgao_acompanhamento=self._extract_orgao_acompanhamento(acompanhamento),
            status=status,
            data_assinatura=None,
            prazo_final=prazo_final,
            periodicidade_relatorio_meses=self._infer_periodicidade_relatorio(
                observacoes=observacoes,
                data_ultimo_relatorio=data_ultimo_relatorio,
                data_proximo_relatorio=data_proximo_relatorio,
            ),
            data_ultimo_relatorio=data_ultimo_relatorio,
            data_proximo_relatorio=data_proximo_relatorio,
            area_m2=area_m2,
            numero_mudas_previsto=self._parse_numero_mudas(row_data.get("numero_mudas_previsto"), area_m2=area_m2),
            servicos_exigidos=servicos_exigidos,
            responsavel_execucao="",
            observacoes=observacoes,
            mpsp_relacionado=self._normalize_yes_no(row_data.get("mpsp_relacionado")),
            inquerito_civil=self._extract_inquerito_civil(servicos_exigidos, observacoes),
            eventos=self._build_eventos(
                data_ultimo_relatorio=data_ultimo_relatorio,
                data_proximo_relatorio=data_proximo_relatorio,
                status=status,
                observacoes=observacoes,
            ),
        )

    def _merge_records(self, existing: Tcra, imported: Tcra) -> Tcra:
        existing_status = normalize_status_label(existing.status)
        imported_status = normalize_status_label(imported.status)
        resolved_status = imported_status or existing_status
        if existing_status in {"Cumprido", "Arquivado"} and imported_status not in {"Cumprido", "Arquivado"}:
            resolved_status = existing_status

        return Tcra(
            uid=existing.uid,
            numero_processo=self._prefer_text(imported.numero_processo, existing.numero_processo),
            numero_tcra=self._prefer_text(imported.numero_tcra, existing.numero_tcra),
            local=self._prefer_text(imported.local, existing.local),
            endereco=self._prefer_text(imported.endereco, existing.endereco),
            bairro=self._prefer_text(imported.bairro, existing.bairro),
            orgao_acompanhamento=self._prefer_text(imported.orgao_acompanhamento, existing.orgao_acompanhamento),
            status=resolved_status,
            data_assinatura=imported.data_assinatura or existing.data_assinatura,
            prazo_final=imported.prazo_final or existing.prazo_final,
            periodicidade_relatorio_meses=(
                imported.periodicidade_relatorio_meses
                if imported.periodicidade_relatorio_meses is not None
                else existing.periodicidade_relatorio_meses
            ),
            data_ultimo_relatorio=imported.data_ultimo_relatorio or existing.data_ultimo_relatorio,
            data_proximo_relatorio=imported.data_proximo_relatorio or existing.data_proximo_relatorio,
            area_m2=imported.area_m2 if imported.area_m2 is not None else existing.area_m2,
            numero_mudas_previsto=(
                imported.numero_mudas_previsto
                if imported.numero_mudas_previsto is not None
                else existing.numero_mudas_previsto
            ),
            servicos_exigidos=self._merge_text_block(existing.servicos_exigidos, imported.servicos_exigidos),
            responsavel_execucao=self._prefer_text(imported.responsavel_execucao, existing.responsavel_execucao),
            observacoes=self._merge_text_block(existing.observacoes, imported.observacoes),
            mpsp_relacionado=self._merge_yes_no(existing.mpsp_relacionado, imported.mpsp_relacionado),
            inquerito_civil=self._merge_text_block(existing.inquerito_civil, imported.inquerito_civil),
            eventos=self._merge_eventos(existing.eventos, imported.eventos),
        )

    @staticmethod
    def _prefer_text(imported_value: object, existing_value: object) -> str:
        imported_text = _stringify(imported_value)
        return imported_text or _stringify(existing_value)

    @staticmethod
    def _merge_text_block(existing_value: object, imported_value: object) -> str:
        existing_text = _stringify(existing_value)
        imported_text = _stringify(imported_value)
        if not existing_text:
            return imported_text
        if not imported_text:
            return existing_text
        if imported_text.casefold() == existing_text.casefold():
            return existing_text
        if imported_text.casefold() in existing_text.casefold():
            return existing_text
        if existing_text.casefold() in imported_text.casefold():
            return imported_text
        return f"{existing_text}\n{imported_text}"

    @staticmethod
    def _merge_yes_no(existing_value: object, imported_value: object) -> str:
        existing_text = _stringify(existing_value)
        imported_text = _stringify(imported_value)
        if imported_text.casefold() in {"sim", "s"} or existing_text.casefold() in {"sim", "s"}:
            return "Sim"
        return imported_text or existing_text

    @staticmethod
    def _merge_eventos(existing_eventos: Sequence[TcraEvento], imported_eventos: Sequence[TcraEvento]) -> list[TcraEvento]:
        merged = list(existing_eventos)
        seen = {
            (
                evento.data_evento,
                _stringify(evento.tipo_evento).casefold(),
                _stringify(evento.descricao).casefold(),
                evento.prazo_resultante,
                _stringify(evento.status_resultante).casefold(),
            )
            for evento in existing_eventos
        }
        for evento in imported_eventos:
            signature = (
                evento.data_evento,
                _stringify(evento.tipo_evento).casefold(),
                _stringify(evento.descricao).casefold(),
                evento.prazo_resultante,
                _stringify(evento.status_resultante).casefold(),
            )
            if signature in seen:
                continue
            seen.add(signature)
            merged.append(evento)
        merged.sort(
            key=lambda evento: (
                evento.data_evento or date.min,
                _stringify(evento.tipo_evento).casefold(),
                _stringify(evento.descricao).casefold(),
                evento.prazo_resultante or date.max,
            )
        )
        return [
            TcraEvento(
                sequence=index,
                data_evento=evento.data_evento,
                tipo_evento=_stringify(evento.tipo_evento),
                descricao=_stringify(evento.descricao),
                prazo_resultante=evento.prazo_resultante,
                status_resultante=_stringify(evento.status_resultante),
            )
            for index, evento in enumerate(merged, start=1)
        ]

    def _parse_date(self, value: object) -> date | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value

        text = _stringify(value)
        if not text or text == "-":
            return None
        text = text.replace(" 00:00:00", "")
        for candidate in (text, text.split(" ")[0]):
            try:
                return date.fromisoformat(candidate)
            except ValueError:
                pass
        for fmt in ("%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        match = re.search(r"(\d{2}/\d{2}/\d{2,4})", text)
        if match:
            found = match.group(1)
            for fmt in ("%d/%m/%Y", "%d/%m/%y"):
                try:
                    return datetime.strptime(found, fmt).date()
                except ValueError:
                    continue
        return None

    def _parse_area(self, value: object) -> float | None:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = _stringify(value)
        if not text or text == "-":
            return None
        if text.startswith("="):
            return self._evaluate_formula_number(text)
        try:
            if "," in text and "." in text:
                return float(text.replace(".", "").replace(",", "."))
            return float(text.replace(",", "."))
        except ValueError:
            return None

    def _parse_numero_mudas(self, value: object, *, area_m2: float | None) -> int | None:
        if value is None:
            if area_m2 is None:
                return None
            return max(int(math.floor(area_m2 / 6.0)), 0)
        if isinstance(value, (int, float)):
            numeric = float(value)
            if numeric < 0:
                return None
            return int(numeric)

        text = _stringify(value)
        if not text or text == "-":
            return None
        if self._looks_like_coordinate(text):
            return None
        if text.startswith("=") and area_m2 is not None and "ROUNDDOWN" in text.upper():
            return max(int(math.floor(area_m2 / 6.0)), 0)
        if text.startswith("="):
            evaluated = self._evaluate_formula_number(text)
            if evaluated is None:
                return None
            return int(evaluated)
        try:
            return int(float(text.replace(",", ".")))
        except ValueError:
            return None

    def _evaluate_formula_number(self, formula: str) -> float | None:
        clean = _stringify(formula)
        if not clean.startswith("="):
            return None
        expression = clean[1:].strip()
        if re.fullmatch(r"[0-9+\-*/().,\s]+", expression):
            expression = expression.replace(",", ".")
            try:
                tree = ast.parse(expression, mode="eval")
                return float(_SafeNumericExpression().visit(tree))
            except Exception:
                return None
        return None

    @staticmethod
    def _normalize_note_text(value: object) -> str:
        text = _stringify(value)
        return text.lstrip("* ").strip()

    @staticmethod
    def _normalize_yes_no(value: object) -> str:
        text = _normalize_header(value)
        if text in {"sim", "s"}:
            return "Sim"
        if text in {"nao", "n"}:
            return "Nao"
        return _stringify(value)

    @staticmethod
    def _looks_like_coordinate(text: str) -> bool:
        normalized = text.replace(",", ".")
        return bool(re.fullmatch(r"-?\d{1,3}\.\d+", normalized))

    @staticmethod
    def _extract_bairro(endereco: str) -> str:
        parts = [part.strip() for part in endereco.split(" - ") if part.strip()]
        if len(parts) >= 2:
            return parts[-1]
        return ""

    @staticmethod
    def _extract_orgao_acompanhamento(acompanhamento: str) -> str:
        text = acompanhamento.strip()
        if not text or TcraExcelService._looks_like_coordinate(text):
            return ""
        normalized = text.casefold()
        if any(keyword in normalized for keyword in ("cumprid", "arquivad", "sem validade", "vencid")):
            return ""
        return text

    def _infer_status(
        self,
        *,
        acompanhamento: str,
        observacoes: str,
        servicos: str,
        prazo_final: date | None,
        data_proximo_relatorio: date | None,
    ) -> str:
        combined = " ".join(part for part in (acompanhamento, observacoes, servicos) if part).casefold()
        if "sem validade" in combined:
            return "Sem validade"
        if "cumprid" in combined:
            return "Cumprido"
        if "arquivad" in combined:
            return "Arquivado"
        if prazo_final is not None and prazo_final < self.today:
            return "Prazo vencido"
        if data_proximo_relatorio is not None and data_proximo_relatorio < self.today:
            return "Relatório pendente"
        if acompanhamento.strip():
            return "Em acompanhamento"
        return ""

    def _infer_periodicidade_relatorio(
        self,
        *,
        observacoes: str,
        data_ultimo_relatorio: date | None,
        data_proximo_relatorio: date | None,
    ) -> int | None:
        text = observacoes.casefold()
        anos = re.search(r"cada\s+(\d+)\s+anos?", text)
        if anos:
            return int(anos.group(1)) * 12
        meses = re.search(r"cada\s+(\d+)\s+mes", text)
        if meses:
            return int(meses.group(1))
        if data_ultimo_relatorio and data_proximo_relatorio and data_proximo_relatorio >= data_ultimo_relatorio:
            month_delta = (data_proximo_relatorio.year - data_ultimo_relatorio.year) * 12 + (
                data_proximo_relatorio.month - data_ultimo_relatorio.month
            )
            if month_delta > 0:
                return month_delta
        return None

    @staticmethod
    def _extract_inquerito_civil(servicos: str, observacoes: str) -> str:
        for text in (servicos, observacoes):
            if "inquerito civil" in _normalize_header(text):
                return text
        return ""

    def _build_eventos(
        self,
        *,
        data_ultimo_relatorio: date | None,
        data_proximo_relatorio: date | None,
        status: str,
        observacoes: str,
    ) -> list[TcraEvento]:
        if data_ultimo_relatorio is not None:
            return [
                TcraEvento(
                    sequence=1,
                    data_evento=data_ultimo_relatorio,
                    tipo_evento="Relatório",
                    descricao="Último relatório importado da planilha legada.",
                    prazo_resultante=data_proximo_relatorio,
                    status_resultante=status,
                )
            ]
        if observacoes:
            return [
                TcraEvento(
                    sequence=1,
                    data_evento=None,
                    tipo_evento="Observacao",
                    descricao=observacoes,
                    prazo_resultante=data_proximo_relatorio,
                    status_resultante=status,
                )
            ]
        return []
