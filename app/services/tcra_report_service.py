from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from functools import partial
from typing import Any, Sequence
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models.tcra import Tcra
from app.services.report_service import _build_pdf_header, _draw_pdf_page_frame
from app.services.report_service_support import build_report_metadata_rows
from app.services.tcra_insights_service import build_sla_summary, build_workload_snapshot
from app.services.tcra_records_service import (
    AGENDA_SCOPE_30D,
    AGENDA_SCOPE_7D,
    AGENDA_SCOPE_PENDENTES,
    AGENDA_SCOPE_VENCIDOS,
    STATUS_CUMPRIDO,
    build_work_agenda,
    build_quality_queue,
    build_record_overview,
    compute_metrics,
    resolve_operational_status,
    suggest_issue_fix,
    tcra_is_mpsp_related,
)


TCRA_PDF_REPORT_TITLE = "Relat\u00f3rio Operacional de TCRAs"
TCRA_PDF_SOURCE_LABEL = "Base operacional de TCRAs"


@dataclass(frozen=True)
class TcraPdfExportOptions:
    include_summary: bool = True
    include_current_records: bool = True
    include_upcoming_reports: bool = True
    include_quality_queue: bool = False
    include_critical_agenda: bool = True
    include_agenda_7d: bool = False
    include_agenda_30d: bool = False
    include_inbox: bool = False

    @classmethod
    def empty_selection(cls) -> "TcraPdfExportOptions":
        return cls(
            include_summary=False,
            include_current_records=False,
            include_upcoming_reports=False,
            include_quality_queue=False,
            include_critical_agenda=False,
            include_agenda_7d=False,
            include_agenda_30d=False,
            include_inbox=False,
        )

    def has_any_section(self) -> bool:
        return any(
            (
                self.include_summary,
                self.include_current_records,
                self.include_upcoming_reports,
                self.include_quality_queue,
                self.include_critical_agenda,
                self.include_agenda_7d,
                self.include_agenda_30d,
                self.include_inbox,
            )
        )


HEADER_FILL = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FDE7E9", end_color="FDE7E9", fill_type="solid")
DONE_FILL = PatternFill(start_color="E3F3E7", end_color="E3F3E7", fill_type="solid")
SOON_FILL = PatternFill(start_color="FFF4D6", end_color="FFF4D6", fill_type="solid")
BORDER_STYLE = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
FONT_BOLD = Font(bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


def _stringify(value: object) -> str:
    return str(value or "").strip()


def _format_date(value: date | None) -> str:
    if value is None:
        return ""
    return value.strftime("%d/%m/%Y")


def _build_summary_rows(records: Sequence[Tcra], *, today: date | None = None) -> list[tuple[str, Any]]:
    metrics = compute_metrics(records, today=today)
    overview = build_record_overview(records, today=today)
    sla_summary = build_sla_summary(records, today=today)
    workload_snapshot = build_workload_snapshot(records, today=today)
    workload_text = "--"
    if workload_snapshot.entries:
        top = workload_snapshot.entries[0]
        workload_text = f"{top.responsavel} ({top.total_count} termo(s), score {top.workload_score})"
    return [
        ("Total de TCRAs", metrics["count_total"]),
        ("Ativos", metrics["count_ativos"]),
        ("Cumpridos", metrics["count_cumpridos"]),
        ("Alertas", metrics["count_alertas"]),
        ("Prazo vencido", metrics["count_prazo_vencido"]),
        ("Relatório pendente", metrics["count_relatorio_pendente"]),
        ("Próximo relatório em 30 dias", metrics["count_relatorio_proximo_30d"]),
        ("Relacionados ao MPSP", metrics["count_mpsp_relacionados"]),
        ("Sem número TCRA", metrics["count_sem_numero_tcra"]),
        ("Sem responsável", metrics["count_sem_responsavel"]),
        ("Prazo interno atrasado", sla_summary.overdue_count),
        ("Prazo interno escalado", sla_summary.escalated_count),
        ("Maior carga", workload_text),
        ("Com eventos", overview.com_eventos_count),
    ]


def _build_data_rows(records: Sequence[Tcra], *, today: date | None = None) -> list[list[str]]:
    rows: list[list[str]] = []
    for record in records:
        rows.append(
            [
                _stringify(record.numero_processo),
                _stringify(record.numero_tcra),
                _stringify(record.local),
                _stringify(record.endereco),
                _stringify(record.bairro),
                _stringify(record.orgao_acompanhamento),
                resolve_operational_status(record, today=today),
                _format_date(record.prazo_final),
                _format_date(record.data_ultimo_relatorio),
                _format_date(record.data_proximo_relatorio),
                "" if record.area_m2 is None else str(record.area_m2),
                "" if record.numero_mudas_previsto is None else str(record.numero_mudas_previsto),
                _stringify(record.responsavel_execucao),
                "Sim" if tcra_is_mpsp_related(record) else "Não",
                _stringify(record.inquerito_civil),
                _stringify(record.servicos_exigidos),
                _stringify(record.observacoes),
                " | ".join(
                    filter(
                        None,
                        (
                            " | ".join(
                                part
                                for part in (
                                    f"{_format_date(evento.data_evento)} {evento.tipo_evento}: {evento.descricao}".strip(),
                                    f"Protocolo {getattr(evento, 'protocolo', '')}".strip()
                                    if getattr(evento, "protocolo", "")
                                    else "",
                                    f"Doc {getattr(evento, 'documento_ref', '')}".strip()
                                    if getattr(evento, "documento_ref", "")
                                    else "",
                                )
                                if part
                            )
                            for evento in record.eventos
                        ),
                    )
                ),
            ]
        )
    return rows


def _agenda_section_rows(records: Sequence[Tcra], *, scope: str, today: date | None = None, limit: int = 8) -> tuple[tuple[str, str, str, str], ...]:
    return tuple(
        (
            item.prioridade_label,
            item.termo_label,
            item.local,
            item.detalhe,
        )
        for item in build_work_agenda(records, scope=scope, today=today, limit=limit)
    )


def _resolve_pdf_col_widths(total_width: float, weights: Sequence[float]) -> list[float]:
    normalized = [max(float(weight), 0.0) for weight in weights]
    total_weight = sum(normalized) or float(len(normalized) or 1)
    return [total_width * (weight / total_weight) for weight in normalized]


def _pdf_text(value: object, *, empty_placeholder: str = "--") -> str:
    text = _stringify(value)
    if not text:
        text = empty_placeholder
    return escape(text).replace("\n", "<br/>")


def _pdf_row(
    values: Sequence[object],
    *,
    style: ParagraphStyle,
    centered_columns: set[int] | None = None,
) -> list[Paragraph]:
    centered_columns = centered_columns or set()
    row: list[Paragraph] = []
    for index, value in enumerate(values):
        if index in centered_columns:
            centered_style = ParagraphStyle(
                f"{style.name}_center_{index}",
                parent=style,
                alignment=1,
            )
            row.append(Paragraph(_pdf_text(value), centered_style))
        else:
            row.append(Paragraph(_pdf_text(value), style))
    return row


def _build_pdf_table(
    *,
    headers: Sequence[str],
    rows: Sequence[Sequence[object]],
    total_width: float,
    column_weights: Sequence[float],
    header_style: ParagraphStyle,
    cell_style: ParagraphStyle,
    header_background: colors.Color,
    centered_columns: Sequence[int] = (),
) -> Table:
    centered = set(centered_columns)
    table_rows = [
        [Paragraph(_pdf_text(header, empty_placeholder=""), header_style) for header in headers]
    ]
    table_rows.extend(
        _pdf_row(row, style=cell_style, centered_columns=centered)
        for row in rows
    )
    table = Table(
        table_rows,
        repeatRows=1,
        colWidths=_resolve_pdf_col_widths(total_width, column_weights),
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), header_background),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("WORDWRAP", (0, 0), (-1, -1), "CJK"),
            ]
        )
    )
    return table


def _build_tcra_pdf_elements(
    records: Sequence[Tcra],
    *,
    filter_summary: str,
    content_width: float,
    styles,
    today: date | None = None,
    options: TcraPdfExportOptions | None = None,
):
    export_options = options or TcraPdfExportOptions()
    normal_style = styles["Normal"]
    normal_style.fontSize = 8
    table_header_style = ParagraphStyle(
        "TcraPdfHeaderResponsive",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7,
        textColor=colors.whitesmoke,
        alignment=1,
        leading=8,
    )
    table_cell_style = ParagraphStyle(
        "TcraPdfCellResponsive",
        parent=styles["Normal"],
        fontSize=7,
        leading=9,
        alignment=0,
    )
    section_title = ParagraphStyle(
        "TcraPdfSectionTitle",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#143A5A"),
        spaceAfter=6,
    )
    metadata_rows = build_report_metadata_rows(
        filter_summary,
        source_label=TCRA_PDF_SOURCE_LABEL,
    )
    elements = list(_build_pdf_header(styles, title=TCRA_PDF_REPORT_TITLE, metadata_rows=metadata_rows))

    if export_options.include_summary:
        elements.extend(
            [
                Paragraph("Resumo do Relat\u00f3rio", section_title),
                _build_pdf_table(
                    headers=["Indicador", "Valor"],
                    rows=[[label, str(value)] for label, value in _build_summary_rows(records, today=today)],
                    total_width=min(content_width * 0.50, 360),
                    column_weights=[0.68, 0.32],
                    header_style=table_header_style,
                    cell_style=table_cell_style,
                    header_background=colors.HexColor("#1F4E78"),
                    centered_columns=[1],
                ),
                Spacer(1, 0.15 * inch),
            ]
        )

    overview = build_record_overview(records, today=today)
    if export_options.include_upcoming_reports and overview.upcoming_reports:
        elements.extend(
            [
                Paragraph("Pr\u00f3ximos relat\u00f3rios", section_title),
                _build_pdf_table(
                    headers=["Termo", "Data"],
                    rows=[
                        [
                            _stringify(sample.numero_tcra or sample.numero_processo or sample.local),
                            _format_date(sample.data_proximo_relatorio),
                        ]
                        for sample in overview.upcoming_reports
                    ],
                    total_width=min(content_width * 0.48, 360),
                    column_weights=[0.72, 0.28],
                    header_style=table_header_style,
                    cell_style=table_cell_style,
                    header_background=colors.HexColor("#4F81BD"),
                    centered_columns=[1],
                ),
                Spacer(1, 0.15 * inch),
            ]
        )

    quality_queue = build_quality_queue(records, today=today, limit=6)
    if export_options.include_quality_queue and quality_queue:
        elements.extend(
            [
                Paragraph("Qualidade cadastral", section_title),
                _build_pdf_table(
                    headers=["Severidade", "Termo", "Local", "Revis\u00e3o", "Sugest\u00e3o"],
                    rows=[
                        [
                            item.severity_label,
                            item.termo_label,
                            item.local,
                            item.detalhe,
                            suggest_issue_fix(item.issues[0]) if item.issues else "",
                        ]
                        for item in quality_queue
                    ],
                    total_width=content_width,
                    column_weights=[0.12, 0.14, 0.20, 0.27, 0.27],
                    header_style=table_header_style,
                    cell_style=table_cell_style,
                    header_background=colors.HexColor("#8B3D3D"),
                ),
                Spacer(1, 0.15 * inch),
            ]
        )

    agenda_sections = []
    if export_options.include_critical_agenda:
        agenda_sections.append(("Pend\u00eancias cr\u00edticas", AGENDA_SCOPE_VENCIDOS, colors.HexColor("#4F81BD")))
    if export_options.include_agenda_7d:
        agenda_sections.append(("Agenda de trabalho - 7 dias", AGENDA_SCOPE_7D, colors.HexColor("#6C8EAD")))
    if export_options.include_agenda_30d:
        agenda_sections.append(("Agenda de trabalho - 30 dias", AGENDA_SCOPE_30D, colors.HexColor("#4F81BD")))
    if export_options.include_inbox:
        agenda_sections.append(("Inbox operacional", AGENDA_SCOPE_PENDENTES, colors.HexColor("#3E6488")))

    for title, scope, header_color in agenda_sections:
        rows = _agenda_section_rows(records, scope=scope, today=today, limit=6)
        if not rows:
            continue
        elements.extend(
            [
                Paragraph(title, section_title),
                _build_pdf_table(
                    headers=["Prioridade", "Termo", "Local", "A\u00e7\u00e3o"],
                    rows=[list(row) for row in rows],
                    total_width=content_width,
                    column_weights=[0.12, 0.15, 0.23, 0.50],
                    header_style=table_header_style,
                    cell_style=table_cell_style,
                    header_background=header_color,
                ),
                Spacer(1, 0.15 * inch),
            ]
        )

    if export_options.include_current_records:
        record_rows = [
            [
                _stringify(record.numero_processo),
                _stringify(record.numero_tcra),
                _stringify(record.local),
                resolve_operational_status(record, today=today),
                _format_date(record.prazo_final),
                _format_date(record.data_proximo_relatorio),
                _stringify(record.orgao_acompanhamento),
                _stringify(record.responsavel_execucao),
            ]
            for record in records
        ]
        elements.extend(
            [
                Paragraph("Recorte atual de TCRAs", section_title),
                _build_pdf_table(
                    headers=[
                        "Processo",
                        "TCRA",
                        "Local",
                        "Status",
                        "Prazo",
                        "Pr\u00f3x. rel.",
                        "\u00d3rg\u00e3o",
                        "Resp.",
                    ],
                    rows=record_rows,
                    total_width=content_width,
                    column_weights=[0.11, 0.10, 0.25, 0.13, 0.09, 0.10, 0.10, 0.12],
                    header_style=table_header_style,
                    cell_style=table_cell_style,
                    header_background=colors.HexColor("#1F4E78"),
                ),
            ]
        )

    return elements


def _append_labeled_section(
    worksheet,
    *,
    start_row: int,
    title: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
) -> int:
    worksheet.cell(row=start_row, column=1, value=title).font = FONT_BOLD
    header_row = start_row + 1
    for column_index, value in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=column_index, value=value)
        cell.font = FONT_BOLD
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER

    current_row = header_row + 1
    if rows:
        for row_values in rows:
            for column_index, value in enumerate(row_values, start=1):
                worksheet.cell(row=current_row, column=column_index, value=value)
            current_row += 1
    else:
        worksheet.cell(row=current_row, column=1, value="Nenhum item no recorte atual.")
        current_row += 1
    return current_row + 1


def _style_worksheet(worksheet) -> None:
    for row in worksheet.iter_rows():
        row_fill = None
        if row and row[0].row > 1:
            status_value = _stringify(row[6].value if len(row) > 6 else "")
            if status_value == STATUS_CUMPRIDO:
                row_fill = DONE_FILL
            elif status_value == "Prazo vencido":
                row_fill = ALERT_FILL
            elif status_value == "Relatório pendente":
                row_fill = SOON_FILL

        for cell in row:
            cell.border = BORDER_STYLE
            if cell.row == 1:
                cell.font = FONT_BOLD
                cell.fill = HEADER_FILL
                cell.alignment = ALIGN_CENTER
            else:
                if row_fill is not None:
                    cell.fill = row_fill
                if len(_stringify(cell.value)) <= 20:
                    cell.alignment = ALIGN_CENTER
                else:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

    for column_cells in worksheet.columns:
        max_length = max(len(_stringify(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 42)

    worksheet.auto_filter.ref = worksheet.dimensions


def export_tcra_excel_report(
    path: str,
    records: Sequence[Tcra],
    *,
    filter_summary: str,
    today: date | None = None,
) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Resumo"
    data_sheet = workbook.create_sheet("TCRAs")

    summary_sheet["A1"] = "Relatório Operacional de TCRAs"
    summary_sheet["A1"].font = Font(bold=True, size=14)
    summary_sheet["A2"] = "Gerado em"
    summary_sheet["B2"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    summary_sheet["A3"] = "Filtros"
    summary_sheet["B3"] = filter_summary

    summary_rows = _build_summary_rows(records, today=today)
    summary_sheet["A5"] = "Indicador"
    summary_sheet["B5"] = "Valor"
    for cell in ("A5", "B5"):
        summary_sheet[cell].font = FONT_BOLD
        summary_sheet[cell].fill = HEADER_FILL
        summary_sheet[cell].alignment = ALIGN_CENTER
    for row_index, (label, value) in enumerate(summary_rows, start=6):
        summary_sheet.cell(row=row_index, column=1, value=label)
        summary_sheet.cell(row=row_index, column=2, value=value)

    overview = build_record_overview(records, today=today)
    next_row = len(summary_rows) + 8
    next_row = _append_labeled_section(
        summary_sheet,
        start_row=next_row,
        title="Próximos relatórios",
        headers=("Termo", "Data"),
        rows=tuple(
            (
                _stringify(sample.numero_tcra or sample.numero_processo or sample.local),
                _format_date(sample.data_proximo_relatorio),
            )
            for sample in overview.upcoming_reports
        ),
    )
    next_row = _append_labeled_section(
        summary_sheet,
        start_row=next_row,
        title="Qualidade cadastral",
        headers=("Severidade", "Termo", "Local", "Revisao", "Sugestao"),
        rows=tuple(
            (
                item.severity_label,
                item.termo_label,
                item.local,
                item.detalhe,
                suggest_issue_fix(item.issues[0]) if item.issues else "",
            )
            for item in build_quality_queue(records, today=today, limit=8)
        ),
    )
    next_row = _append_labeled_section(
        summary_sheet,
        start_row=next_row,
        title="Pendencias criticas",
        headers=("Prioridade", "Termo", "Local", "Acao"),
        rows=_agenda_section_rows(records, scope=AGENDA_SCOPE_VENCIDOS, today=today, limit=8),
    )
    next_row = _append_labeled_section(
        summary_sheet,
        start_row=next_row,
        title="Agenda de trabalho - 7 dias",
        headers=("Prioridade", "Termo", "Local", "Acao"),
        rows=_agenda_section_rows(records, scope=AGENDA_SCOPE_7D, today=today, limit=8),
    )
    next_row = _append_labeled_section(
        summary_sheet,
        start_row=next_row,
        title="Agenda de trabalho - 30 dias",
        headers=("Prioridade", "Termo", "Local", "Acao"),
        rows=_agenda_section_rows(records, scope=AGENDA_SCOPE_30D, today=today, limit=8),
    )
    _append_labeled_section(
        summary_sheet,
        start_row=next_row,
        title="Inbox operacional",
        headers=("Prioridade", "Termo", "Local", "Acao"),
        rows=_agenda_section_rows(records, scope=AGENDA_SCOPE_PENDENTES, today=today, limit=8),
    )

    headers = [
        "Processo",
        "Numero TCRA",
        "Local",
        "Endereço",
        "Bairro",
        "Órgão",
        "Status operacional",
        "Prazo final",
        "Último relatório",
        "Próximo relatório",
        "Area (m2)",
        "Numero de mudas",
        "Responsavel",
        "MPSP",
        "Inquerito civil",
        "Servicos exigidos",
        "Observações",
        "Eventos",
    ]
    data_sheet.append(headers)
    for row in _build_data_rows(records, today=today):
        data_sheet.append(row)

    _style_worksheet(summary_sheet)
    _style_worksheet(data_sheet)
    workbook.save(path)


def _export_tcra_pdf_report_legacy(
    path: str,
    records: Sequence[Tcra],
    *,
    filter_summary: str,
    today: date | None = None,
) -> None:
    document = SimpleDocTemplate(
        path,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=18,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.fontSize = 14
    normal_style = styles["Normal"]
    normal_style.fontSize = 8
    table_header_style = ParagraphStyle(
        "TcraPdfHeader",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7,
        textColor=colors.whitesmoke,
        alignment=1,
        leading=8,
    )
    table_cell_style = ParagraphStyle(
        "TcraPdfCell",
        parent=styles["Normal"],
        fontSize=7,
        leading=8,
        alignment=0,
    )

    elements = [
        Paragraph("Relatório Operacional de TCRAs", title_style),
        Spacer(1, 8),
        Paragraph(f"<b>Filtros:</b> {filter_summary}", normal_style),
        Spacer(1, 8),
    ]

    summary_table = Table(
        [["Indicador", "Valor"]] + [[label, str(value)] for label, value in _build_summary_rows(records, today=today)],
        colWidths=[220, 120],
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 10))

    overview = build_record_overview(records, today=today)
    if overview.upcoming_reports:
        upcoming_table = Table(
            [["Termo", "Data"]] + [
                [
                    _stringify(sample.numero_tcra or sample.numero_processo or sample.local),
                    _format_date(sample.data_proximo_relatorio),
                ]
                for sample in overview.upcoming_reports
            ],
            colWidths=[250, 110],
        )
        upcoming_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        elements.append(Paragraph("<b>Próximos relatórios</b>", normal_style))
        elements.append(upcoming_table)
        elements.append(Spacer(1, 10))

    quality_queue = build_quality_queue(records, today=today, limit=6)
    if quality_queue:
        quality_table = Table(
            [["Severidade", "Termo", "Local", "Revisao", "Sugestao"]]
            + [
                [
                    item.severity_label,
                    item.termo_label,
                    item.local,
                    item.detalhe,
                    suggest_issue_fix(item.issues[0]) if item.issues else "",
                ]
                for item in quality_queue
            ],
            colWidths=[70, 90, 135, 180, 170],
        )
        quality_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8B3D3D")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        elements.append(Paragraph("<b>Qualidade cadastral</b>", normal_style))
        elements.append(quality_table)
        elements.append(Spacer(1, 10))

    critical_rows = _agenda_section_rows(records, scope=AGENDA_SCOPE_VENCIDOS, today=today, limit=6)
    if critical_rows:
        agenda_table = Table(
            [["Prioridade", "Termo", "Local", "Acao"]]
            + [list(row) for row in critical_rows],
            colWidths=[85, 100, 160, 240],
        )
        agenda_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        elements.append(Paragraph("<b>Pendencias criticas</b>", normal_style))
        elements.append(agenda_table)
        elements.append(Spacer(1, 10))

    agenda_7d_rows = _agenda_section_rows(records, scope=AGENDA_SCOPE_7D, today=today, limit=6)
    if agenda_7d_rows:
        agenda_7d_table = Table(
            [["Prioridade", "Termo", "Local", "Acao"]] + [list(row) for row in agenda_7d_rows],
            colWidths=[85, 100, 160, 240],
        )
        agenda_7d_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6C8EAD")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        elements.append(Paragraph("<b>Agenda de trabalho - 7 dias</b>", normal_style))
        elements.append(agenda_7d_table)
        elements.append(Spacer(1, 10))

    agenda_30d_rows = _agenda_section_rows(records, scope=AGENDA_SCOPE_30D, today=today, limit=6)
    if agenda_30d_rows:
        agenda_30d_table = Table(
            [["Prioridade", "Termo", "Local", "Acao"]] + [list(row) for row in agenda_30d_rows],
            colWidths=[85, 100, 160, 240],
        )
        agenda_30d_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        elements.append(Paragraph("<b>Agenda de trabalho - 30 dias</b>", normal_style))
        elements.append(agenda_30d_table)
        elements.append(Spacer(1, 10))

    pending_rows = _agenda_section_rows(records, scope=AGENDA_SCOPE_PENDENTES, today=today, limit=6)
    if pending_rows:
        pending_table = Table(
            [["Prioridade", "Termo", "Local", "Acao"]] + [list(row) for row in pending_rows],
            colWidths=[85, 100, 160, 240],
        )
        pending_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3E6488")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        elements.append(Paragraph("<b>Inbox operacional</b>", normal_style))
        elements.append(pending_table)
        elements.append(Spacer(1, 10))

    headers = [
        "Processo",
        "TCRA",
        "Local",
        "Status",
        "Prazo",
        "Prox. rel.",
        "Órgão",
        "Resp.",
    ]
    table_rows = [[Paragraph(label, table_header_style) for label in headers]]
    for record in records:
        row = [
            _stringify(record.numero_processo),
            _stringify(record.numero_tcra),
            _stringify(record.local),
            resolve_operational_status(record, today=today),
            _format_date(record.prazo_final),
            _format_date(record.data_proximo_relatorio),
            _stringify(record.orgao_acompanhamento),
            _stringify(record.responsavel_execucao),
        ]
        table_rows.append([Paragraph(value or "--", table_cell_style) for value in row])

    records_table = Table(
        table_rows,
        repeatRows=1,
        colWidths=[90, 85, 180, 90, 65, 70, 70, 90],
    )
    records_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
            ]
        )
    )
    elements.append(Paragraph("<b>Recorte atual de TCRAs</b>", normal_style))
    elements.append(records_table)
    document.build(elements)


def export_tcra_pdf_report(
    path: str,
    records: Sequence[Tcra],
    *,
    filter_summary: str,
    today: date | None = None,
    options: TcraPdfExportOptions | None = None,
    emitted_by: str = "",
) -> None:
    export_options = options or TcraPdfExportOptions()
    if not export_options.has_any_section():
        raise ValueError("Selecione ao menos um bloco para exportar no PDF de TCRAs.")

    document = SimpleDocTemplate(
        path,
        pagesize=landscape(A4),
        rightMargin=22,
        leftMargin=22,
        topMargin=22,
        bottomMargin=26,
    )
    generated_label = datetime.now().strftime("%d/%m/%Y %H:%M")
    elements = _build_tcra_pdf_elements(
        records,
        filter_summary=filter_summary,
        content_width=document.width,
        styles=getSampleStyleSheet(),
        today=today,
        options=export_options,
    )
    document.build(
        elements,
        onFirstPage=partial(
            _draw_pdf_page_frame,
            title=TCRA_PDF_REPORT_TITLE,
            generated_label=generated_label,
            emitted_by=emitted_by,
        ),
        onLaterPages=partial(
            _draw_pdf_page_frame,
            title=TCRA_PDF_REPORT_TITLE,
            generated_label=generated_label,
            emitted_by=emitted_by,
        ),
    )
