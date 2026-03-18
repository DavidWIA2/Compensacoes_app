from pathlib import Path
import pytest
import openpyxl

from app.models.compensacao import Compensacao
from app.services.excel_service import BACKUP_FOLDER_NAME, ExcelService
from app.services.records_service import remove_accents


SHEET_NAME = "Compensa\u00e7\u00f5es"


def build_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(
        [
            "Of\u00edcio/ Processo",
            "Eletr\u00f4nico",
            "Caixa",
            "Av. Tec.",
            "Compensa\u00e7\u00e3o",
            "Endere\u00e7o",
            "Microbacia",
            "Compensado",
        ]
    )
    ws.append(["123/2026", "SIM", "CX-1", "AT-1", 8, "Rua A", "Gregorio", ""])
    wb.save(path)


def build_legacy_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(
        [
            "Oficio/ Processo",
            "Eletronico",
            "Caixa",
            "Av. Tec.",
            "Compensacao",
            "Endereco",
            "Microbacia",
            "Compensado",
        ]
    )
    ws.append(["123/2026", "SIM", "CX-1", "AT-1", 8, "Rua A", "Gregorio", ""])
    wb.save(path)


def make_record(**overrides) -> Compensacao:
    base = {
        "excel_row": 3,
        "oficio_processo": "456/2026",
        "eletronico": "NAO",
        "caixa": "CX-2",
        "av_tec": "AT-2",
        "compensacao": "12",
        "endereco": "Rua B",
        "microbacia": "Monjolinho",
        "compensado": "SIM",
        "latitude": "-22.01",
        "longitude": "-47.89",
    }
    base.update(overrides)
    return Compensacao(**base)


def test_load_exposes_lat_lon_headers_and_migrates_workbook(tmp_path):
    path = tmp_path / "compensacoes.xlsx"
    build_workbook(path)

    service = ExcelService()
    records = service.load(str(path))

    assert len(records) == 1
    # Na nova estrutura, Latitude e Longitude estao nas colunas 12 e 13
    assert service.ws.cell(row=1, column=12).value == "Latitude"
    assert service.ws.cell(row=1, column=13).value == "Longitude"

    # O servico muta o workbook em memoria, mas nao deve salvar no disco durante o load 
    # a menos que tenha gerado UIDs. Como nosso build_workbook nao tem UID, ele vai gerar e salvar.
    persisted = openpyxl.load_workbook(path)
    ws = persisted[SHEET_NAME]
    # Se gerou UID, ele salvou. E o load do ExcelService FRESH gera UID se faltar.
    assert ws.cell(row=1, column=14).value == "UID"
    assert ws.cell(row=1, column=12).value == "Latitude"


def test_load_does_not_create_backup(tmp_path):
    path = tmp_path / "compensacoes.xlsx"
    build_workbook(path)

    service = ExcelService()
    service.load(str(path))

    assert not (tmp_path / BACKUP_FOLDER_NAME).exists()


def test_load_does_not_duplicate_legacy_headers_without_accents(tmp_path):
    path = tmp_path / "compensacoes_legado.xlsx"
    build_legacy_workbook(path)

    service = ExcelService()
    service.load(str(path))

    persisted = openpyxl.load_workbook(path)
    ws = persisted[SHEET_NAME]
    headers = [cell.value for cell in ws[1]]
    normalized = [remove_accents(str(header)).strip().upper() for header in headers if header]

    assert len(headers) == 14
    assert normalized.count("OFICIO/ PROCESSO") == 1
    assert normalized.count("ELETRONICO") == 1
    assert normalized.count("COMPENSACAO") == 1
    assert normalized.count("ENDERECO") == 1


def test_add_new_persists_record_and_creates_backup(tmp_path):
    path = tmp_path / "compensacoes.xlsx"
    build_workbook(path)

    service = ExcelService()
    service.load(str(path))

    new_row = service.add_new(make_record())

    reloaded = openpyxl.load_workbook(path)
    ws = reloaded[SHEET_NAME]

    assert new_row == 3
    assert ws.cell(row=1, column=12).value == "Latitude"
    assert ws.cell(row=1, column=13).value == "Longitude"
    assert ws.cell(row=3, column=1).value == "456/2026"
    assert ws.cell(row=3, column=12).value == "-22.01"
    assert ws.cell(row=3, column=13).value == "-47.89"
    assert (tmp_path / BACKUP_FOLDER_NAME).exists()


def test_add_new_skips_partially_filled_rows(tmp_path):
    path = tmp_path / "compensacoes_parcial.xlsx"
    build_legacy_workbook(path)

    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET_NAME]
    ws.append(["", "NAO", "CX-2", "AT-2", 4, "Rua B", "Medeiros", ""])
    wb.save(path)

    service = ExcelService()
    service.load(str(path))

    new_row = service.add_new(make_record(av_tec="AT-3", oficio_processo="789/2026"))

    reloaded = openpyxl.load_workbook(path)
    ws = reloaded[SHEET_NAME]

    assert new_row == 4
    assert ws.cell(row=3, column=4).value == "AT-2"
    assert ws.cell(row=4, column=1).value == "789/2026"
    assert ws.cell(row=4, column=4).value == "AT-3"


def test_delete_record_shift_up_removes_row(tmp_path):
    path = tmp_path / "compensacoes.xlsx"
    build_workbook(path)

    service = ExcelService()
    service.load(str(path))
    service.add_new(make_record())

    service.delete_record_shift_up(2)

    reloaded = openpyxl.load_workbook(path)
    ws = reloaded[SHEET_NAME]

    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == "456/2026"


def test_delete_record_shift_up_raises_when_uid_is_missing(tmp_path):
    path = tmp_path / "compensacoes_uid.xlsx"
    build_workbook(path)

    service = ExcelService()
    records = service.load(str(path))

    with pytest.raises(LookupError, match="UID"):
        service.delete_record_shift_up(records[0].excel_row, records[0].uid + "-missing")

    reloaded = openpyxl.load_workbook(path)
    ws = reloaded[SHEET_NAME]
    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == "123/2026"


def test_save_workbook_cleans_temp_file_when_replace_fails(tmp_path, monkeypatch):
    path = tmp_path / "compensacoes_locked.xlsx"
    build_workbook(path)

    service = ExcelService()
    records = service.load(str(path))
    records[0].caixa = "CX-ALTERADA"

    monkeypatch.setattr("app.services.excel_service.os.replace", lambda src, dst: (_ for _ in ()).throw(PermissionError("locked")))

    with pytest.raises(PermissionError):
        service.save_edit(records[0])

    temp_files = [candidate for candidate in tmp_path.glob("compensacoes_*.xlsx") if candidate.name != path.name]
    assert temp_files == []

    reloaded = openpyxl.load_workbook(path)
    ws = reloaded[SHEET_NAME]
    assert ws.cell(row=2, column=3).value == "CX-1"


def test_read_all_requires_loaded_path():
    service = ExcelService()

    with pytest.raises(ValueError, match="Nenhum arquivo Excel carregado"):
        service.read_all()
