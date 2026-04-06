from datetime import date

import openpyxl

from app.models.tcra import Tcra
from app.models.tcra_evento import TcraEvento
from app.services.tcra_excel_service import TCRA_SHEET_NAME, TcraExcelService
from app.services.tcra_sqlite_service import TcraSqliteService


def build_legacy_tcra_workbook(path, *, sheet_name=TCRA_SHEET_NAME):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(
        [
            "Processo",
            "Local",
            "Endereco",
            "Relat. Periodico",
            "Ultimo Rel.",
            "Prazo",
            "Servicos a realizar",
            "Tamanho",
            "No de Mudas",
            "Acompanhamento",
            "",
            "MPSP?",
        ]
    )
    worksheet.append(
        [
            "26207/2019",
            "Sistema de Lazer - Residencial Itamarati",
            "Rua Ireneu Couto - Residencial Itamarati",
            date(2025, 3, 10),
            date(2024, 4, 11),
            date(2026, 4, 1),
            "Tratos Culturais regulares antes do prazo",
            2920,
            "=ROUNDDOWN(H2/6,0)",
            "CETESB",
            "*Relatorio a cada 5 anos",
            "Nao",
        ]
    )
    worksheet.append(
        [
            "2360/2021",
            "Varjao",
            "Margem da Rod. Eng. Thales de Lorena Peixoto Junior",
            "-",
            date(2025, 1, 3),
            "-",
            "Inquerito Civil Arquivado em 23/01/2025",
            12577,
            "=ROUNDDOWN(H3/6,0)",
            "Cumprido",
            "*Cumprido",
            "Sim",
        ]
    )
    worksheet.append([None, None, None, None, None, None, None, 1400, "=ROUNDDOWN(H4/6,0)", None, None, None])
    workbook.save(path)


def build_problematic_tcra_workbook(path):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = TCRA_SHEET_NAME
    worksheet.append(
        [
            "Processo",
            "Local",
            "Endereco",
            "Relat. Periodico",
            "Ultimo Rel.",
            "Prazo",
            "Servicos a realizar",
            "Tamanho",
            "No de Mudas",
            "Acompanhamento",
            "",
            "MPSP?",
        ]
    )
    worksheet.append(
        [
            None,
            "Area verde central",
            "Rua A - Centro",
            date(2025, 1, 1),
            date(2025, 2, 1),
            date(2026, 12, 31),
            "Tratos culturais",
            1200,
            "-22.0171",
            "-47.8891",
            "*Aguardando definicao",
            "",
        ]
    )
    workbook.save(path)


def test_tcra_excel_service_loads_legacy_sheet_with_operational_mapping(tmp_path):
    workbook_path = tmp_path / "tcras.xlsx"
    build_legacy_tcra_workbook(workbook_path)

    service = TcraExcelService(sqlite_service=TcraSqliteService(db_path=tmp_path / "local.db"), today=date(2026, 4, 3))
    tcras = service.load_workbook(workbook_path)

    assert len(tcras) == 2

    first = tcras[0]
    assert first.numero_processo == "26207/2019"
    assert first.local == "Sistema de Lazer - Residencial Itamarati"
    assert first.bairro == "Residencial Itamarati"
    assert first.orgao_acompanhamento == "CETESB"
    assert first.status == "Prazo vencido"
    assert first.periodicidade_relatorio_meses == 60
    assert first.data_ultimo_relatorio == date(2024, 4, 11)
    assert first.data_proximo_relatorio == date(2025, 3, 10)
    assert first.area_m2 == 2920.0
    assert first.numero_mudas_previsto == 486
    assert first.mpsp_relacionado == "Nao"
    assert first.eventos[0].tipo_evento == "Relatorio"

    second = tcras[1]
    assert second.status == "Cumprido"
    assert second.inquerito_civil == "Inquerito Civil Arquivado em 23/01/2025"
    assert second.numero_mudas_previsto == 2096
    assert second.mpsp_relacionado == "Sim"


def test_tcra_excel_service_analyzes_workbook_with_import_issues(tmp_path):
    workbook_path = tmp_path / "tcras-problematicos.xlsx"
    build_problematic_tcra_workbook(workbook_path)

    service = TcraExcelService(sqlite_service=TcraSqliteService(db_path=tmp_path / "local.db"), today=date(2026, 4, 3))
    analysis = service.analyze_workbook(workbook_path)
    issue_codes = {issue.code for issue in analysis.issues}

    assert analysis.importable_count == 1
    assert analysis.skipped_count == 0
    assert analysis.missing_columns == ()
    assert {"processo_ausente", "numero_mudas_invalido", "acompanhamento_deslocado", "datas_relatorio_inconsistentes"} <= issue_codes


def test_tcra_excel_service_imports_workbook_into_local_database(tmp_path):
    workbook_path = tmp_path / "tcras.xlsx"
    build_legacy_tcra_workbook(workbook_path)
    sqlite_service = TcraSqliteService(db_path=tmp_path / "local.db")
    service = TcraExcelService(sqlite_service=sqlite_service, today=date(2026, 4, 3))

    imported = service.import_workbook(workbook_path)
    persisted = sqlite_service.list_tcras()

    assert imported == 2
    assert len(persisted) == 2
    assert persisted[0].numero_processo == "2360/2021"
    assert persisted[0].eventos[0].status_resultante == "Cumprido"
    assert persisted[1].numero_mudas_previsto == 486


def test_tcra_excel_service_merge_workbook_preserves_manual_fields_and_updates_matches(tmp_path):
    workbook_path = tmp_path / "tcras.xlsx"
    build_legacy_tcra_workbook(workbook_path)
    sqlite_service = TcraSqliteService(db_path=tmp_path / "local.db")
    sqlite_service.upsert_tcra(
        Tcra(
            uid="tcra-manual-1",
            numero_processo="26207/2019",
            numero_tcra="TCRA-MANUAL-001",
            local="Sistema de Lazer - Residencial Itamarati",
            endereco="Rua Ireneu Couto - Residencial Itamarati",
            bairro="Residencial Itamarati",
            orgao_acompanhamento="CETESB",
            status="Em acompanhamento",
            data_assinatura=None,
            prazo_final=None,
            periodicidade_relatorio_meses=None,
            data_ultimo_relatorio=None,
            data_proximo_relatorio=None,
            area_m2=None,
            numero_mudas_previsto=None,
            servicos_exigidos="Monitoramento fotografico manual",
            responsavel_execucao="Equipe interna",
            observacoes="Observacao manual",
            mpsp_relacionado="Nao",
            inquerito_civil="",
            eventos=[
                TcraEvento(
                    sequence=1,
                    data_evento=date(2026, 1, 10),
                    tipo_evento="Despacho",
                    descricao="Evento manual previo",
                    prazo_resultante=None,
                    status_resultante="Em acompanhamento",
                )
            ],
        )
    )
    service = TcraExcelService(sqlite_service=sqlite_service, today=date(2026, 4, 3))

    result = service.merge_workbook(workbook_path)
    persisted = sqlite_service.find_tcra_by_uid("tcra-manual-1")

    assert result.created_count == 1
    assert result.updated_count == 1
    assert persisted is not None
    assert persisted.numero_tcra == "TCRA-MANUAL-001"
    assert persisted.responsavel_execucao == "Equipe interna"
    assert "Observacao manual" in persisted.observacoes
    assert persisted.prazo_final == date(2026, 4, 1)
    assert len(persisted.eventos) >= 2


def test_tcra_excel_service_analysis_reports_skipped_partial_rows(tmp_path):
    workbook_path = tmp_path / "tcras.xlsx"
    build_legacy_tcra_workbook(workbook_path)

    service = TcraExcelService(sqlite_service=TcraSqliteService(db_path=tmp_path / "local.db"), today=date(2026, 4, 3))
    analysis = service.analyze_workbook(workbook_path)

    assert analysis.importable_count == 2
    assert analysis.skipped_count == 1
    assert any(issue.code == "linha_descartada_sem_identificacao" for issue in analysis.issues)


def test_tcra_excel_service_analysis_summary_lines_include_top_issue_codes(tmp_path):
    workbook_path = tmp_path / "tcras-problematicos.xlsx"
    build_problematic_tcra_workbook(workbook_path)

    service = TcraExcelService(sqlite_service=TcraSqliteService(db_path=tmp_path / "local.db"), today=date(2026, 4, 3))
    analysis = service.analyze_workbook(workbook_path)
    summary_lines = analysis.summary_lines()

    assert summary_lines[0] == "TCRAs importaveis: 1"
    assert any("Principais ocorrencias:" in line for line in summary_lines)
    assert any("Primeiros termos:" in line for line in summary_lines)


def test_tcra_excel_service_accepts_legacy_apostrophe_sheet_name(tmp_path):
    workbook_path = tmp_path / "tcras-apostrofo.xlsx"
    build_legacy_tcra_workbook(workbook_path, sheet_name="TCRA's")

    service = TcraExcelService(sqlite_service=TcraSqliteService(db_path=tmp_path / "local.db"), today=date(2026, 4, 3))
    analysis = service.analyze_workbook(workbook_path)

    assert analysis.worksheet_name == "TCRA's"
    assert analysis.importable_count == 2
