from datetime import date, datetime
from pathlib import Path
from types import SimpleNamespace

import openpyxl
from openpyxl import load_workbook

from app.application.use_cases.tcra_module_operations import TcraModuleOperations
from app.models.tcra import Tcra
from app.models.tcra_evento import TcraEvento
from app.services.access_service import AccessEnvironment, AppAccessSession
from app.services.tcra_excel_service import TCRA_SHEET_NAME
from app.services.tcra_sqlite_service import TcraSqliteService


def make_tcra(**overrides) -> Tcra:
    base = {
        "uid": "tcra-1",
        "numero_processo": "26207/2019",
        "numero_tcra": "TCRA-2019-001",
        "local": "Sistema de Lazer - Residencial Itamarati",
        "endereco": "Rua Ireneu Couto",
        "bairro": "Residencial Itamarati",
        "orgao_acompanhamento": "CETESB",
        "status": "Em acompanhamento",
        "data_assinatura": date(2019, 6, 1),
        "prazo_final": date(2026, 4, 1),
        "periodicidade_relatorio_meses": 60,
        "data_ultimo_relatorio": date(2024, 4, 11),
        "data_proximo_relatorio": date(2025, 3, 10),
        "area_m2": 2920.0,
        "numero_mudas_previsto": 486,
        "servicos_exigidos": "Tratos culturais regulares",
        "responsavel_execucao": "Secretaria Municipal",
        "observacoes": "Relatorio a cada 5 anos",
        "mpsp_relacionado": "Nao",
        "inquerito_civil": "",
        "eventos": [
            TcraEvento(
                sequence=1,
                data_evento=date(2024, 4, 11),
                tipo_evento="Relatorio",
                descricao="Relatorio periodico protocolado",
                prazo_resultante=date(2025, 3, 10),
                status_resultante="Em acompanhamento",
            )
        ],
    }
    base.update(overrides)
    return Tcra(**base)


def build_legacy_tcra_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = TCRA_SHEET_NAME
    worksheet.append(
        [
            "Processo",
            "Local",
            "Endereco",
            "Relat. Periodico",
            "Ultimo Rel.",
            "Prazo",
            "Servicos a realizar",
            "Tamanho",
            "No de Mudas",
            "Acompanhamento",
            "",
            "MPSP?",
        ]
    )
    worksheet.append(
        [
            "26207/2019",
            "Sistema de Lazer - Residencial Itamarati",
            "Rua Ireneu Couto - Residencial Itamarati",
            date(2025, 3, 10),
            date(2024, 4, 11),
            date(2026, 4, 1),
            "Tratos Culturais regulares antes do prazo",
            2920,
            "=ROUNDDOWN(H2/6,0)",
            "CETESB",
            "*Relatorio a cada 5 anos",
            "Nao",
        ]
    )
    worksheet.append(
        [
            "2360/2021",
            "Varjao",
            "Margem da Rod. Eng. Thales de Lorena Peixoto Junior",
            "-",
            date(2025, 1, 3),
            "-",
            "Inquerito Civil Arquivado em 23/01/2025",
            12577,
            "=ROUNDDOWN(H3/6,0)",
            "Cumprido",
            "*Cumprido",
            "Sim",
        ]
    )
    workbook.save(path)


class FakeRemoteTcraRpcService:
    def __init__(self, *, save_uid: str = "remote-tcra-1"):
        self.save_uid = save_uid
        self.save_calls = []
        self.delete_calls = []
        self.save_records_calls = []

    def save_record(self, client, **kwargs):
        self.save_calls.append({"client": client, **kwargs})
        return SimpleNamespace(uid=self.save_uid or getattr(kwargs["record"], "uid", ""), tcra_count=1)

    def delete_record(self, client, **kwargs):
        self.delete_calls.append({"client": client, **kwargs})
        return SimpleNamespace(uid=kwargs["uid"], tcra_count=0)

    def save_records(self, client, **kwargs):
        self.save_records_calls.append({"client": client, **kwargs})
        return SimpleNamespace(uid="", tcra_count=len(kwargs["records"]), imported_count=len(kwargs["records"]))


class FakeRemoteTcraSyncService:
    def __init__(self, *, synced_tcras=None, fail: bool = False):
        self.synced_tcras = list(synced_tcras or [])
        self.fail = fail
        self.calls = []

    def sync_authenticated_client(self, client, *, local_db_path=None, session_path=None):
        self.calls.append({"client": client, "local_db_path": str(local_db_path or ""), "session_path": session_path})
        if self.fail:
            raise RuntimeError("remote tcra cache offline")
        TcraSqliteService(db_path=local_db_path).replace_all(self.synced_tcras)
        return SimpleNamespace(
            local_db_path=str(local_db_path or ""),
            session_path=session_path,
            tcra_count=len(self.synced_tcras),
        )


class FakeAccessService:
    def __init__(self, *, client=None, sync_service=None):
        self.client = client or object()
        self.production_sync_service = sync_service
        self.calls = []

    def create_authenticated_client(self, access_session):
        self.calls.append(access_session)
        return self.client


def make_production_session(**overrides) -> AppAccessSession:
    base = {
        "environment": AccessEnvironment.PRODUCTION,
        "label": "Producao",
        "auth_mode": "password",
        "user_id": "user-123",
        "user_email": "analista@prefeitura.sp.gov.br",
        "supabase_url": "https://yonvcnnkewzoqwnnmcdx.supabase.co",
        "local_db_path": "C:/tmp/producao.db",
        "local_session_path": "session://banco-local",
        "app_role": "editor",
        "access_token": "token",
        "refresh_token": "refresh-token",
    }
    base.update(overrides)
    return AppAccessSession(**base)


def make_operations(
    service: TcraSqliteService,
    audit_calls: list[dict] | None = None,
    *,
    access_session=None,
    access_service=None,
    remote_tcra_service=None,
) -> TcraModuleOperations:
    audit_calls = audit_calls if audit_calls is not None else []
    audit_service = SimpleNamespace(append_session_event=lambda **kwargs: audit_calls.append(kwargs))
    return TcraModuleOperations(
        service,
        today=date(2026, 4, 3),
        audit_service_provider=lambda: audit_service,
        session_path_provider=lambda: "session://banco-local",
        access_session_provider=lambda: access_session,
        access_service=access_service,
        remote_tcra_service=remote_tcra_service,
    )


def test_tcra_module_operations_load_save_delete_and_dashboard(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    audit_calls: list[dict] = []
    operations = make_operations(service, audit_calls)

    load_result = operations.load_records()
    assert load_result.records == ()
    assert load_result.search_index == {}

    save_result = operations.save_record(
        make_tcra(uid="", numero_processo="777/2026", numero_tcra="TCRA-2026-777", local="Parque Linear"),
        pending_audit_metadata={"event_change_action": "add", "event_change_type": "Vistoria"},
    )

    assert save_result.status == "saved"
    assert save_result.saved_uid
    assert save_result.saved_record is not None
    assert audit_calls[-1]["action"] == "TCRA_CREATE"
    assert audit_calls[-1]["metadata"]["event_change_type"] == "Vistoria"

    load_result = operations.load_records()
    assert load_result.records[0].uid == save_result.saved_uid
    assert save_result.saved_uid in load_result.search_index

    dashboard_payload = operations.build_dashboard_payload(load_result.records)
    assert dashboard_payload.overview is not None
    assert len(dashboard_payload.agenda_items) >= 1

    delete_result = operations.delete_record(save_result.saved_uid)
    assert delete_result.status == "deleted"
    assert service.list_tcras() == []
    assert audit_calls[-1]["action"] == "TCRA_DELETE"


def test_tcra_module_operations_returns_duplicate_and_invalid_results(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all([make_tcra(uid="tcra-1")])
    operations = make_operations(service)

    duplicate_result = operations.save_record(
        make_tcra(uid="", numero_tcra="TCRA-2019-001", numero_processo="999/2026", local="Novo local")
    )
    assert duplicate_result.status == "duplicate"
    assert duplicate_result.duplicate_record is not None

    invalid_result = operations.save_record(
        make_tcra(
            uid="",
            numero_tcra="",
            numero_processo="888/2026",
            local="Area Norte",
            status="Cumprido",
            data_ultimo_relatorio=date(2026, 4, 10),
            data_proximo_relatorio=date(2026, 4, 1),
        )
    )
    assert invalid_result.status == "invalid"
    assert "Próximo relatório não pode ser anterior ao último relatório." in invalid_result.consistency_issues


def test_tcra_module_operations_applies_bulk_event_action(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(uid="tcra-1", data_proximo_relatorio=date(2026, 4, 1)),
            make_tcra(uid="tcra-2", numero_tcra="TCRA-2", numero_processo="7205/2014", data_proximo_relatorio=date(2026, 4, 2)),
        ]
    )
    audit_calls: list[dict] = []
    operations = make_operations(service, audit_calls)

    result = operations.apply_bulk_action(
        service.list_tcras(),
        {
            "action": "evento",
            "event_preset": "cumprimento",
            "event_date": "10/04/2026",
            "event_deadline": "10/04/2026",
        },
        parse_date=lambda text, _label: datetime.strptime(text, "%d/%m/%Y").date(),
        event_presets=(
            {
                "key": "cumprimento",
                "tipo_evento": "Cumprimento",
                "descricao": "Termo cumprido.",
                "status_resultante": "Cumprido",
            },
        ),
    )

    assert result.action == "evento"
    assert len(result.updated_uids) == 2
    updated_records = service.get_tcras_by_uids(result.updated_uids)
    assert all(record.status == "Cumprido" for record in updated_records)
    assert all(record.data_proximo_relatorio is None for record in updated_records)
    assert audit_calls[-1]["action"] == "TCRA_BULK_UPDATE"


def test_tcra_module_operations_imports_and_exports_reports(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    audit_calls: list[dict] = []
    operations = make_operations(service, audit_calls)
    workbook_path = tmp_path / "tcras.xlsx"
    build_legacy_tcra_workbook(workbook_path)

    analysis = operations.analyze_import_workbook(workbook_path)
    assert analysis.importable_count == 2

    import_result = operations.execute_import_merge(analysis)
    assert import_result.merge_result.importable_count == 2
    assert import_result.preferred_uid
    assert audit_calls[-1]["action"] == "TCRA_IMPORT"

    records = service.list_tcras()
    excel_path = tmp_path / "tcra-export.xlsx"
    pdf_path = tmp_path / "tcra-export.pdf"
    operations.export_excel_report(str(excel_path), records, filter_summary="Todos")
    operations.export_pdf_report(str(pdf_path), records, filter_summary="Todos")

    workbook = load_workbook(excel_path)
    assert workbook.sheetnames == ["Resumo", "TCRAs"]
    assert pdf_path.exists() is True
    assert pdf_path.stat().st_size > 0


def test_tcra_module_operations_uses_remote_save_in_production(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    remote_record = make_tcra(uid="remote-tcra-1", numero_tcra="TCRA-REMOTE")
    remote_sync = FakeRemoteTcraSyncService(synced_tcras=[remote_record])
    access_service = FakeAccessService(sync_service=remote_sync)
    remote_rpc = FakeRemoteTcraRpcService(save_uid="remote-tcra-1")
    audit_calls: list[dict] = []
    operations = make_operations(
        service,
        audit_calls,
        access_session=make_production_session(),
        access_service=access_service,
        remote_tcra_service=remote_rpc,
    )

    result = operations.save_record(make_tcra(uid="", numero_tcra="TCRA-REMOTE"))

    assert result.status == "saved"
    assert result.authority_source == "remote"
    assert result.saved_uid == "remote-tcra-1"
    assert result.sync_issues == ()
    assert len(remote_rpc.save_calls) == 1
    assert remote_rpc.save_calls[0]["action"] == "TCRA_CREATE"
    assert len(remote_sync.calls) == 1
    assert service.get_tcra("remote-tcra-1") is not None
    assert audit_calls == []


def test_tcra_module_operations_refreshes_remote_snapshot_on_demand_in_production(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all([make_tcra(uid="stale-tcra", numero_tcra="TCRA-STALE")])
    remote_record = make_tcra(uid="remote-tcra", numero_tcra="TCRA-REMOTE")
    remote_sync = FakeRemoteTcraSyncService(synced_tcras=[remote_record])
    access_service = FakeAccessService(sync_service=remote_sync)
    operations = make_operations(
        service,
        access_session=make_production_session(),
        access_service=access_service,
        remote_tcra_service=FakeRemoteTcraRpcService(),
    )

    result = operations.load_records(refresh_remote=True)

    assert len(remote_sync.calls) == 1
    assert remote_sync.calls[0]["session_path"] == "session://banco-local"
    assert [record.uid for record in result.records] == ["remote-tcra"]
    assert result.sync_issues == ()


def test_tcra_module_operations_keeps_local_cache_when_remote_refresh_fails(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all([make_tcra(uid="cached-tcra", numero_tcra="TCRA-CACHE")])
    remote_sync = FakeRemoteTcraSyncService(synced_tcras=[make_tcra(uid="remote-tcra")], fail=True)
    operations = make_operations(
        service,
        access_session=make_production_session(),
        access_service=FakeAccessService(sync_service=remote_sync),
        remote_tcra_service=FakeRemoteTcraRpcService(),
    )

    result = operations.load_records(refresh_remote=True)

    assert len(remote_sync.calls) == 1
    assert [record.uid for record in result.records] == ["cached-tcra"]
    assert "snapshot remoto de TCRA" in " | ".join(result.sync_issues)


def test_tcra_module_operations_does_not_refresh_remote_snapshot_in_local_mode(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all([make_tcra(uid="local-tcra", numero_tcra="TCRA-LOCAL")])
    remote_sync = FakeRemoteTcraSyncService(synced_tcras=[make_tcra(uid="remote-tcra")])
    operations = make_operations(
        service,
        access_service=FakeAccessService(sync_service=remote_sync),
        remote_tcra_service=FakeRemoteTcraRpcService(),
    )

    result = operations.load_records(refresh_remote=True)

    assert remote_sync.calls == []
    assert [record.uid for record in result.records] == ["local-tcra"]


def test_tcra_module_operations_remote_save_falls_back_to_local_cache_when_refresh_fails(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    remote_sync = FakeRemoteTcraSyncService(fail=True)
    access_service = FakeAccessService(sync_service=remote_sync)
    remote_rpc = FakeRemoteTcraRpcService(save_uid="remote-tcra-1")
    operations = make_operations(
        service,
        access_session=make_production_session(),
        access_service=access_service,
        remote_tcra_service=remote_rpc,
    )

    result = operations.save_record(make_tcra(uid="", numero_tcra="TCRA-FALLBACK"))

    assert result.authority_source == "remote"
    assert "cache local de TCRA" in " | ".join(result.sync_issues)
    assert service.get_tcra("remote-tcra-1") is not None


def test_tcra_module_operations_uses_remote_bulk_and_import_in_production(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    existing = make_tcra(uid="tcra-1", numero_tcra="TCRA-1", numero_processo="26207/2019")
    service.replace_all([existing])
    remote_sync = FakeRemoteTcraSyncService(synced_tcras=[make_tcra(uid="tcra-1", status="Cumprido")])
    access_service = FakeAccessService(sync_service=remote_sync)
    remote_rpc = FakeRemoteTcraRpcService(save_uid="tcra-1")
    operations = make_operations(
        service,
        access_session=make_production_session(),
        access_service=access_service,
        remote_tcra_service=remote_rpc,
    )

    bulk_result = operations.apply_bulk_action(
        service.list_tcras(),
        {"action": "status", "status": "Cumprido"},
        parse_date=lambda _text, _label: None,
        event_presets=(),
    )

    workbook_path = tmp_path / "tcras.xlsx"
    build_legacy_tcra_workbook(workbook_path)
    analysis = operations.analyze_import_workbook(workbook_path)
    import_result = operations.execute_import_merge(analysis)

    assert bulk_result.authority_source == "remote"
    assert bulk_result.updated_uids == ("tcra-1",)
    assert import_result.authority_source == "remote"
    assert len(remote_rpc.save_records_calls) == 2
    assert remote_rpc.save_records_calls[0]["action"] == "TCRA_BULK_UPDATE"
    assert remote_rpc.save_records_calls[1]["action"] == "TCRA_IMPORT"
