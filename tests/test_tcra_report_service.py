from datetime import date

from openpyxl import load_workbook
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph

from app.models.tcra import Tcra
from app.models.tcra_evento import TcraEvento
from app.services.tcra_report_service import (
    TcraPdfExportOptions,
    _build_pdf_table,
    _build_tcra_pdf_elements,
    export_tcra_excel_report,
    export_tcra_pdf_report,
)


def make_tcra(**overrides) -> Tcra:
    base = {
        "uid": "tcra-1",
        "numero_processo": "26207/2019",
        "numero_tcra": "TCRA-2019-001",
        "local": "Sistema de Lazer",
        "endereco": "Rua Ireneu Couto - Itamarati",
        "bairro": "Itamarati",
        "orgao_acompanhamento": "CETESB",
        "status": "Em acompanhamento",
        "data_assinatura": date(2019, 6, 1),
        "prazo_final": date(2026, 4, 1),
        "periodicidade_relatorio_meses": 60,
        "data_ultimo_relatorio": date(2024, 4, 11),
        "data_proximo_relatorio": date(2026, 4, 20),
        "area_m2": 2920.0,
        "numero_mudas_previsto": 486,
        "servicos_exigidos": "Tratos culturais regulares",
        "responsavel_execucao": "Secretaria Municipal",
        "observacoes": "Relatorio a cada 5 anos",
        "mpsp_relacionado": "Nao",
        "inquerito_civil": "",
        "eventos": [
            TcraEvento(
                sequence=1,
                data_evento=date(2024, 4, 11),
                tipo_evento="Relatorio",
                descricao="Relatorio periodico protocolado",
                prazo_resultante=date(2026, 4, 20),
                status_resultante="Em acompanhamento",
            )
        ],
    }
    base.update(overrides)
    return Tcra(**base)


def test_export_tcra_excel_report_generates_summary_and_data_sheet(tmp_path):
    export_path = tmp_path / "tcra-report.xlsx"
    records = [
        make_tcra(uid="tcra-1"),
        make_tcra(uid="tcra-2", numero_tcra="", responsavel_execucao="", orgao_acompanhamento=""),
        make_tcra(uid="tcra-3", status="Cumprido", data_proximo_relatorio=None),
    ]

    export_tcra_excel_report(
        str(export_path),
        records,
        filter_summary="Busca: nenhuma | Status: Todos",
        today=date(2026, 4, 3),
    )

    workbook = load_workbook(export_path)

    assert workbook.sheetnames == ["Resumo", "TCRAs"]
    assert workbook["Resumo"]["A1"].value == "Relatório Operacional de TCRAs"
    assert workbook["Resumo"]["B3"].value == "Busca: nenhuma | Status: Todos"
    summary_values = {
        str(cell.value)
        for row in workbook["Resumo"].iter_rows()
        for cell in row
        if cell.value is not None
    }
    assert "Qualidade cadastral" in summary_values
    assert "Inbox operacional" in summary_values
    assert "Pendencias criticas" in summary_values
    assert "Agenda de trabalho - 7 dias" in summary_values
    assert "Agenda de trabalho - 30 dias" in summary_values
    assert "Sugestao" in summary_values
    assert "Cadastro" in summary_values
    assert workbook["TCRAs"]["A2"].value == "26207/2019"
    assert workbook["TCRAs"]["G2"].value in {"Prazo vencido", "Em acompanhamento", "Cumprido", "Relatório pendente"}


def test_export_tcra_pdf_report_creates_non_empty_file(tmp_path):
    export_path = tmp_path / "tcra-report.pdf"
    records = [make_tcra(uid="tcra-1"), make_tcra(uid="tcra-2", status="Cumprido", data_proximo_relatorio=None)]

    export_tcra_pdf_report(
        str(export_path),
        records,
        filter_summary="Busca: nenhuma | Status: Todos",
        today=date(2026, 4, 3),
    )

    assert export_path.exists() is True
    assert export_path.stat().st_size > 0


def test_build_tcra_pdf_elements_respects_selected_sections():
    records = [make_tcra(uid="tcra-1")]

    elements = _build_tcra_pdf_elements(
        records,
        filter_summary="Busca: nenhuma | Status: Todos",
        content_width=720,
        styles=getSampleStyleSheet(),
        today=date(2026, 4, 3),
        options=TcraPdfExportOptions(
            include_summary=True,
            include_current_records=True,
            include_upcoming_reports=False,
            include_quality_queue=False,
            include_critical_agenda=False,
            include_agenda_7d=False,
            include_agenda_30d=False,
            include_inbox=False,
        ),
    )

    titles = [element.getPlainText() for element in elements if isinstance(element, Paragraph)]

    assert "Resumo do Relatório" in titles
    assert "Recorte atual de TCRAs" in titles
    assert "Próximos relatórios" not in titles
    assert "Qualidade cadastral" not in titles
    assert "Inbox operacional" not in titles


def test_export_tcra_pdf_report_requires_at_least_one_section(tmp_path):
    export_path = tmp_path / "tcra-report-empty.pdf"

    try:
        export_tcra_pdf_report(
            str(export_path),
            [make_tcra(uid="tcra-1")],
            filter_summary="Busca: nenhuma | Status: Todos",
            today=date(2026, 4, 3),
            options=TcraPdfExportOptions(
                include_summary=False,
                include_current_records=False,
                include_upcoming_reports=False,
                include_quality_queue=False,
                include_critical_agenda=False,
                include_agenda_7d=False,
                include_agenda_30d=False,
                include_inbox=False,
            ),
        )
        raised = False
    except ValueError as exc:
        raised = True
        assert "ao menos um bloco" in str(exc)

    assert raised is True


def test_build_pdf_table_uses_wrappable_paragraph_cells():
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet

    styles = getSampleStyleSheet()
    header_style = ParagraphStyle("header_test", parent=styles["Normal"])
    cell_style = ParagraphStyle("cell_test", parent=styles["Normal"])

    table = _build_pdf_table(
        headers=["Coluna A", "Coluna B"],
        rows=[["Texto muito longo para precisar quebrar linha dentro da célula", "Outro texto longo"]],
        total_width=320,
        column_weights=[0.5, 0.5],
        header_style=header_style,
        cell_style=cell_style,
        header_background=colors.HexColor("#1F4E78"),
    )

    assert all(isinstance(cell, Paragraph) for cell in table._cellvalues[0])
    assert all(isinstance(cell, Paragraph) for cell in table._cellvalues[1])
    assert round(sum(table._colWidths), 5) == 320
