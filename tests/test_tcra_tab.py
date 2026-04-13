import os
from datetime import date
from types import SimpleNamespace

import openpyxl
import pytest
from openpyxl import load_workbook

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PySide6.QtWidgets import QApplication, QWidget

import app.ui.tabs.tcra_tab as tcra_tab_module
from app.models.tcra import Tcra
from app.models.tcra_evento import TcraEvento
from app.services.tcra_excel_service import TCRA_SHEET_NAME, TcraExcelService
from app.services.tcra_sqlite_service import TcraSqliteService
from app.ui.components.date_input import DatePickerLineEdit
from app.ui.components.dialogs import TcraEventoEditorDialog, TcraImportPreviewDialog
from app.ui.tabs.tcra_tab import TcraTab


def get_app():
    return QApplication.instance() or QApplication([])


def make_tcra(**overrides) -> Tcra:
    base = {
        "uid": "tcra-1",
        "numero_processo": "26207/2019",
        "numero_tcra": "TCRA-2019-001",
        "local": "Sistema de Lazer - Residencial Itamarati",
        "endereco": "Rua Ireneu Couto",
        "bairro": "Residencial Itamarati",
        "orgao_acompanhamento": "CETESB",
        "status": "Em acompanhamento",
        "data_assinatura": date(2019, 6, 1),
        "prazo_final": date(2026, 4, 1),
        "periodicidade_relatorio_meses": 60,
        "data_ultimo_relatorio": date(2024, 4, 11),
        "data_proximo_relatorio": date(2025, 3, 10),
        "area_m2": 2920.0,
        "numero_mudas_previsto": 486,
        "servicos_exigidos": "Tratos culturais regulares",
        "responsavel_execucao": "Secretaria Municipal",
        "observacoes": "Relatorio a cada 5 anos",
        "mpsp_relacionado": "Nao",
        "inquerito_civil": "",
        "eventos": [
            TcraEvento(
                sequence=1,
                data_evento=date(2024, 4, 11),
                tipo_evento="Relatorio",
                descricao="Relatorio periodico protocolado",
                prazo_resultante=date(2025, 3, 10),
                status_resultante="Em acompanhamento",
            )
        ],
    }
    base.update(overrides)
    return Tcra(**base)


def build_legacy_tcra_workbook(path):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = TCRA_SHEET_NAME
    worksheet.append(
        [
            "Processo",
            "Local",
            "Endereco",
            "Relat. Periodico",
            "Ultimo Rel.",
            "Prazo",
            "Servicos a realizar",
            "Tamanho",
            "No de Mudas",
            "Acompanhamento",
            "",
            "MPSP?",
        ]
    )
    worksheet.append(
        [
            "26207/2019",
            "Sistema de Lazer - Residencial Itamarati",
            "Rua Ireneu Couto - Residencial Itamarati",
            date(2025, 3, 10),
            date(2024, 4, 11),
            date(2026, 4, 1),
            "Tratos Culturais regulares antes do prazo",
            2920,
            "=ROUNDDOWN(H2/6,0)",
            "CETESB",
            "*Relatorio a cada 5 anos",
            "Nao",
        ]
    )
    worksheet.append(
        [
            "2360/2021",
            "Varjao",
            "Margem da Rod. Eng. Thales de Lorena Peixoto Junior",
            "-",
            date(2025, 1, 3),
            "-",
            "Inquerito Civil Arquivado em 23/01/2025",
            12577,
            "=ROUNDDOWN(H3/6,0)",
            "Cumprido",
            "*Cumprido",
            "Sim",
        ]
    )
    workbook.save(path)


@pytest.fixture(autouse=True)
def app():
    return get_app()


def test_tcra_tab_refreshes_cards_table_and_details(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(uid="tcra-1"),
            make_tcra(
                uid="tcra-2",
                numero_tcra="",
                numero_processo="2360/2021",
                local="Varjao",
                bairro="Varjao",
                orgao_acompanhamento="MPSP",
                status="Cumprido",
                prazo_final=date(2024, 1, 1),
                data_proximo_relatorio=None,
                mpsp_relacionado="Sim",
                eventos=[],
            ),
        ]
    )

    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()

    assert tab.table.rowCount() == 2
    assert tab.table.columnCount() == 11
    assert tab.table.horizontalHeaderItem(0).text() == "Prioridade"
    assert tab.table.horizontalHeaderItem(3).text() == "Status"
    assert tab.table.horizontalHeaderItem(4).text() == "Ult. evento"
    assert tab.table.horizontalHeaderItem(5).text() == "Prox. acao"
    assert tab.table.item(0, 4).text()
    assert tab.table.item(0, 5).text()
    assert tab.card_total.lbl_value.text() == "2"
    assert tab.card_cumpridos.lbl_value.text() == "1"
    assert tab.card_mpsp.lbl_value.text() == "1"
    assert "termos" in tab.lbl_context.text()
    assert "Foco do recorte" in tab.lbl_radar_summary.text()
    assert tab.lbl_radar_summary.isWindow() is False
    assert tab.lbl_data_quality.isWindow() is False
    assert tab.lbl_import_status.isHidden() is True
    assert tab.overview_tabs.tabText(0) == "Seleção"
    assert "Inbox operacional (" in tab.overview_tabs.tabText(1)
    assert "Qualidade cadastral (" in tab.overview_tabs.tabText(2)
    assert tab.agenda_table.rowCount() >= 1
    assert "Hoje:" in tab.lbl_agenda_summary.text()
    assert tab.operational_dialog.isVisible() is False
    assert tab.selection_actions_frame.isHidden() is True

    tab.table.selectRow(1)
    get_app().processEvents()

    assert tab.current_form_uid == ""
    assert tab.selected_uid == "tcra-2"
    assert tab.operational_dialog.isVisible() is False
    assert tab.btn_record_details.isEnabled() is True
    assert "Próxima ação:" in tab.record_details.toPlainText()
    assert "Prazo interno de tratamento:" in tab.record_details.toPlainText()
    assert "Varjao" in tab.record_details.toPlainText()
    assert tab.events_table.rowCount() == 0
    assert "nenhum registro" in tab.lbl_event_spotlight_title.text().lower()
    assert tab.selection_actions_frame.isHidden() is False


def test_tcra_tab_initial_prefetch_loads_when_startup_tab_is_active(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(
                uid="tcra-1",
                prazo_final=date(2027, 4, 1),
                data_proximo_relatorio=None,
            )
        ]
    )
    parent = QWidget()
    tab = TcraTab(parent=parent, sqlite_service=service, today=date(2026, 4, 3))
    parent.tabs = SimpleNamespace(currentWidget=lambda: tab)
    tab._records_loaded = False
    tab._initial_prefetch_pending = True
    tab.all_tcras = []
    tab._set_initial_loading_state()

    tab._prefetch_initial_records()

    assert tab._records_loaded is True
    assert tab.table.rowCount() == 1
    assert tab.card_total.lbl_value.text() == "1"


def test_tcra_tab_startup_deadline_alert_lists_actionable_deadlines(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(
                uid="tcra-overdue",
                numero_tcra="TCRA-OVERDUE",
                prazo_final=date(2026, 4, 1),
                data_proximo_relatorio=None,
            ),
            make_tcra(
                uid="tcra-soon",
                numero_tcra="TCRA-SOON",
                prazo_final=date(2026, 4, 10),
                data_proximo_relatorio=None,
            ),
            make_tcra(
                uid="tcra-done",
                numero_tcra="TCRA-DONE",
                status="Cumprido",
                prazo_final=date(2026, 4, 5),
                data_proximo_relatorio=None,
            ),
        ]
    )

    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()

    alert = tab._build_startup_deadline_alert()

    assert alert is not None
    title, message, has_overdue = alert
    assert title == "Prazos de TCRA"
    assert has_overdue is True
    assert "Prazos vencidos: 1" in message
    assert "TCRA-OVERDUE: 01/04/2026 (vencido há 2 dias)" in message
    assert "Prazos que vencem nos próximos 30 dias: 1" in message
    assert "TCRA-SOON: 10/04/2026 (vence em 7 dias)" in message
    assert "TCRA-DONE" not in message
    assert "registros" not in message.lower()


def test_tcra_tab_startup_deadline_alert_waits_for_main_window_ready(tmp_path, monkeypatch):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(
                uid="tcra-overdue",
                numero_tcra="TCRA-OVERDUE",
                prazo_final=date(2026, 4, 1),
                data_proximo_relatorio=None,
            )
        ]
    )
    parent = QWidget()
    parent._startup_close_guard_active = True
    parent.show()
    get_app().processEvents()
    tab = TcraTab(parent=parent, sqlite_service=service, today=date(2026, 4, 3))
    tab.all_tcras = service.list_tcras()
    captured = []

    monkeypatch.setattr(
        tcra_tab_module.QMessageBox,
        "warning",
        lambda *args, **kwargs: captured.append(args[2] if len(args) > 2 else ""),
    )

    tab._show_startup_deadline_alert()

    assert captured == []
    assert tab._startup_deadline_alert_shown is False

    parent._startup_close_guard_active = False
    tab._show_startup_deadline_alert()

    assert captured
    assert "TCRA-OVERDUE" in captured[-1]
    assert tab._startup_deadline_alert_shown is True


def test_tcra_tab_defers_initial_load_until_event_loop_runs(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all([make_tcra(uid="tcra-1")])

    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))

    assert tab._records_loaded is False
    assert tab.table.rowCount() == 0

    get_app().processEvents()

    assert tab._records_loaded is True
    assert tab.table.rowCount() == 1


def test_tcra_tab_surfaces_event_context_at_top_of_editor(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(
                uid="tcra-1",
                eventos=[
                    TcraEvento(
                        sequence=1,
                        data_evento=date(2024, 4, 11),
                        tipo_evento="Relatorio",
                        descricao="Relatorio protocolado.",
                        prazo_resultante=date(2025, 3, 10),
                        status_resultante="Em acompanhamento",
                        protocolo="SEI-321",
                        documento_ref="C:/docs/relatorio.pdf",
                    )
                ],
            )
        ]
    )

    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()

    tab.table.selectRow(0)
    get_app().processEvents()
    tab.btn_open_selected.click()
    get_app().processEvents()

    assert "Ultimo evento:" in tab.lbl_event_spotlight_title.text()
    assert "SEI-321" in tab.lbl_event_spotlight_meta.text()
    assert tab.editor_tabs.tabText(1).startswith("Eventos")
    assert tab.btn_event_open_latest_document.isEnabled() is True


def test_tcra_tab_compacts_list_header_by_default_but_can_expand_context(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(
                uid="tcra-1",
                numero_tcra="",
                responsavel_execucao="",
                eventos=[],
            )
        ]
    )

    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()

    assert tab.header_kicker.isHidden() is True
    assert tab.header_subtitle.isHidden() is True
    assert tab.summary_details_frame.isHidden() is True
    assert tab.filters_hint.isHidden() is True
    assert "Alertas" in tab.lbl_workspace_digest.text()
    assert "Sem responsavel" in tab.lbl_workspace_digest.text()

    tab.btn_toggle_workspace_context.click()
    get_app().processEvents()

    assert tab.header_kicker.isHidden() is False
    assert tab.header_subtitle.isHidden() is False
    assert tab.summary_details_frame.isHidden() is False
    assert tab.filters_hint.isHidden() is False
    assert tab.btn_toggle_workspace_context.text() == "Menos contexto"


def test_tcra_record_details_dialog_can_register_event_from_consulta(monkeypatch):
    persisted = []

    class EventDialog:
        def __init__(self, *args, **kwargs):
            self.kwargs = kwargs

        def exec(self):
            return True

        def values(self):
            return {
                "preset_key": self.kwargs.get("preset_key", ""),
                "data_evento": "06/04/2026",
                "tipo_evento": "Despacho",
                "descricao": "Cobranca registrada na janela de detalhes.",
                "prazo_resultante": "10/05/2026",
                "status_resultante": "Em acompanhamento",
                "protocolo": "SEI-999",
                "documento_ref": "C:/docs/oficio.pdf",
            }

    monkeypatch.setattr(tcra_tab_module, "TcraEventoEditorDialog", EventDialog)

    dialog = tcra_tab_module.TcraRecordDetailsDialog(
        None,
        record=make_tcra(uid="tcra-1", eventos=[]),
        today=date(2026, 4, 3),
        build_event_from_values=lambda sequence, values: TcraEvento(
            sequence=sequence,
            data_evento=date(2026, 4, 6),
            tipo_evento=values["tipo_evento"],
            descricao=values["descricao"],
            prazo_resultante=date(2026, 5, 10),
            status_resultante=values["status_resultante"],
            protocolo=values["protocolo"],
            documento_ref=values["documento_ref"],
        ),
        apply_event_effects_to_record=lambda record: record,
        persist_record_changes=lambda record, metadata: persisted.append((record, dict(metadata))) or record,
    )

    dialog._add_event_with_preset("despacho")
    get_app().processEvents()

    assert len(persisted) == 1
    assert persisted[0][0].eventos[0].tipo_evento == "Despacho"
    assert persisted[0][1]["event_change_action"] == "add"
    assert dialog.events_table.rowCount() == 1
    assert dialog.tabs.tabText(2).startswith("Eventos (1)")


def test_tcra_tab_handle_tab_activated_schedules_window_fit(tmp_path, monkeypatch):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    main_window = SimpleNamespace()
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    tab.main_window = main_window
    calls = []

    monkeypatch.setattr(tab, "refresh_data", lambda preferred_uid=None, refresh_remote=False: calls.append(("refresh", preferred_uid, refresh_remote)))
    monkeypatch.setattr(
        tcra_tab_module,
        "schedule_window_fit",
        lambda window: calls.append(("fit", window)) or True,
    )

    tab.handle_tab_activated()

    assert calls[0] == ("refresh", "", False)
    assert calls[1] == ("fit", main_window)


def test_tcra_evento_editor_dialog_applies_presets():
    dialog = TcraEventoEditorDialog(None)

    dialog.combo_preset.setCurrentIndex(dialog.combo_preset.findData("cumprimento"))
    get_app().processEvents()

    assert dialog.in_tipo_evento.text() == "Cumprimento"
    assert dialog.in_status_resultante.text() == "Cumprido"
    assert "cumprido" in dialog.in_descricao.toPlainText().lower()
    assert dialog.values()["preset_key"] == "cumprimento"


def test_tcra_import_preview_dialog_filters_visible_issues(tmp_path):
    workbook_path = tmp_path / "tcras.xlsx"
    build_legacy_tcra_workbook(workbook_path)
    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook[TCRA_SHEET_NAME]
    worksheet.append([None, "Area Norte", "Rua A - Centro", date(2025, 1, 1), date(2025, 2, 1), None, "Teste", 1200, "-22.0", "-47.8", "", ""])
    workbook.save(workbook_path)

    service = TcraSqliteService(db_path=tmp_path / "local.db")
    analysis = TcraExcelService(sqlite_service=service, today=date(2026, 4, 3)).analyze_workbook(workbook_path)
    dialog = TcraImportPreviewDialog(None, analysis)

    dialog.filter_severity.setCurrentText("warning")
    dialog.search_input.setText("coordenada")
    get_app().processEvents()

    assert dialog.table.rowCount() >= 1
    assert "Mostrando" in dialog.lbl_visible.text()


def test_tcra_tab_preserves_dirty_form_when_reactivated(tmp_path, monkeypatch):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    calls = []

    monkeypatch.setattr(tab, "refresh_data", lambda preferred_uid=None: calls.append(preferred_uid))

    tab.in_numero_processo.setText("999/2026")
    get_app().processEvents()
    tab.handle_tab_activated()

    assert tab.has_pending_form_changes() is True
    assert tab.lbl_form_state.text() == tab.FORM_DIRTY_TEXT
    assert calls == []


def test_tcra_tab_new_tcra_switches_to_cadastro_workspace_and_keeps_form_visible(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))

    assert tab.workspace_tabs.tabText(tab.workspace_tabs.currentIndex()) == "Lista"

    tab.new_tcra()
    get_app().processEvents()

    assert tab.workspace_tabs.tabText(tab.workspace_tabs.currentIndex()) == "Cadastro"
    assert tab.workspace_tabs.tabText(2) == "Resumo"
    assert tab.editor_tabs.currentIndex() == 0
    assert tab.form_panel_body.isHidden() is False
    assert tab.btn_toggle_form_panel.isChecked() is True
    assert tab.editor_operational_panel.isHidden() is True


def test_tcra_tab_cadastro_workspace_hides_form_toggle_and_keeps_form_open(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))

    tab.new_tcra()
    get_app().processEvents()

    assert tab.form_panel_body.isHidden() is False
    assert tab.btn_toggle_form_panel.isHidden() is True
    assert tab.btn_toggle_form_panel.text() == "Ocultar cadastro"
    assert tab.editor_operational_panel.isHidden() is True


def test_tcra_tab_summary_workspace_exposes_read_only_preview(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))

    assert tab.workspace_tabs.indexOf(tab.summary_workspace_page) == 2
    assert tab.details.isReadOnly() is True


def test_tcra_tab_uses_calendar_date_inputs_in_form(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))

    assert isinstance(tab.in_data_assinatura, DatePickerLineEdit)
    assert isinstance(tab.in_prazo_final, DatePickerLineEdit)
    assert isinstance(tab.in_data_ultimo_relatorio, DatePickerLineEdit)
    assert isinstance(tab.in_data_proximo_relatorio, DatePickerLineEdit)


def test_tcra_tab_main_vertical_containers_do_not_force_window_height(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))

    assert tab.workspace_tabs.minimumHeight() == 0
    assert tab.workspace_tabs.sizePolicy().verticalPolicy().name == "Ignored"
    assert tab.list_splitter.minimumHeight() == 0
    assert tab.list_splitter.sizePolicy().verticalPolicy().name == "Ignored"
    assert tab.form_scroll.minimumHeight() == 0
    assert tab.form_scroll.sizePolicy().verticalPolicy().name == "Ignored"
    assert tab.editor_splitter.minimumHeight() == 0
    assert tab.editor_splitter.sizePolicy().verticalPolicy().name == "Ignored"
    assert tab.editor_tabs.minimumHeight() == 0
    assert tab.editor_tabs.sizePolicy().verticalPolicy().name == "Ignored"


def test_tcra_tab_exposes_clearable_inputs_placeholders_and_tooltips(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))

    assert tab.search_input.isClearButtonEnabled() is True
    assert tab.search_input.toolTip() != ""
    assert tab.in_numero_processo.isClearButtonEnabled() is True
    assert "26207/2019" in tab.in_numero_processo.placeholderText()
    assert tab.in_periodicidade.validator() is not None
    assert tab.in_area_m2.validator() is not None
    assert tab.in_numero_mudas.validator() is not None
    assert tab.btn_save.toolTip() != ""
    assert tab.btn_quick_report.toolTip() != ""
    assert tab.in_servicos.placeholderText() != ""


def test_tcra_tab_summary_actions_navigate_to_operational_views(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(uid="tcra-1", prazo_final=date(2026, 3, 1), data_proximo_relatorio=date(2026, 4, 10)),
            make_tcra(
                uid="tcra-2",
                numero_tcra="",
                numero_processo="2360/2021",
                local="Varjao",
                bairro="Varjao",
                orgao_acompanhamento="",
                responsavel_execucao="",
                prazo_final=date(2026, 8, 1),
                data_proximo_relatorio=date(2026, 4, 20),
                eventos=[],
            ),
        ]
    )

    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    assert tab.operational_dialog.isVisible() is False

    tab._open_inbox_overview()
    get_app().processEvents()
    assert tab.operational_dialog.isVisible() is True
    assert tab.overview_tabs.tabText(tab.overview_tabs.currentIndex()).startswith("Inbox")
    assert tab.agenda_table.columnCount() == 6
    assert tab.agenda_table.horizontalHeaderItem(5).text() == "Ação"

    tab._open_quality_overview()
    get_app().processEvents()
    assert tab.overview_tabs.tabText(tab.overview_tabs.currentIndex()).startswith("Qualidade")
    assert tab.quality_table.columnCount() == 5
    assert tab.quality_table.horizontalHeaderItem(4).text() == "Campos"

    tab.btn_summary_upcoming.click()
    get_app().processEvents()
    assert tab.overview_tabs.tabText(tab.overview_tabs.currentIndex()).startswith("Inbox")
    assert tab.quick_filter_mode == tcra_tab_module.QUICK_FILTER_PROXIMOS


def test_tcra_tab_advanced_filters_are_collapsible_and_count_active_flags(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))

    assert tab.advanced_filters_frame.isHidden() is True
    assert tab.btn_toggle_advanced_filters.text() == "Mais filtros"

    tab.chk_only_mpsp.setChecked(True)
    get_app().processEvents()

    assert "(1)" in tab.btn_toggle_advanced_filters.text()

    tab.btn_toggle_advanced_filters.click()
    get_app().processEvents()

    assert tab.advanced_filters_frame.isHidden() is False
    assert tab.btn_toggle_advanced_filters.text() == "Ocultar filtros"


def test_tcra_tab_operational_dialog_opens_and_closes_from_summary_actions(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all([make_tcra(uid="tcra-1", prazo_final=date(2026, 3, 1), data_proximo_relatorio=date(2026, 4, 10))])

    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()

    assert tab.operational_dialog.isVisible() is False

    tab.btn_summary_inbox.click()
    get_app().processEvents()
    assert tab.operational_dialog.isVisible() is True
    assert tab.lbl_overview_title.text() == "Inbox operacional"
    assert "Inbox operacional" in tab.operational_dialog.windowTitle()

    tab.btn_close_overview.click()
    get_app().processEvents()
    assert tab.operational_dialog.isVisible() is False

    tab.btn_summary_quality.click()
    get_app().processEvents()
    assert tab.operational_dialog.isVisible() is True
    assert tab.lbl_overview_title.text() == "Qualidade cadastral"


def test_tcra_tab_inbox_and_quality_expand_from_compact_preview(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(
                uid=f"tcra-{index}",
                numero_tcra="",
                numero_processo=f"{26000 + index}/2019",
                local=f"Area {index}",
                responsavel_execucao="",
                orgao_acompanhamento="",
                eventos=[],
            )
            for index in range(5)
        ]
    )

    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()

    assert tab.agenda_table.rowCount() == 3
    assert tab.quality_table.rowCount() == 3

    tab.btn_agenda_view_all.click()
    get_app().processEvents()
    assert tab.agenda_table.rowCount() == 5

    tab.btn_quality_view_all.click()
    get_app().processEvents()
    assert tab.quality_table.rowCount() == 5


def test_tcra_tab_agenda_scope_buttons_filter_work_queue(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(uid="tcra-1", prazo_final=date(2026, 3, 20), data_proximo_relatorio=date(2026, 4, 10)),
            make_tcra(uid="tcra-2", prazo_final=date(2026, 8, 1), data_proximo_relatorio=date(2026, 4, 5)),
            make_tcra(
                uid="tcra-3",
                numero_tcra="",
                responsavel_execucao="",
                orgao_acompanhamento="",
                prazo_final=date(2026, 8, 1),
                data_proximo_relatorio=date(2026, 8, 1),
            ),
        ]
    )
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()

    assert tab.agenda_scope == tcra_tab_module.AGENDA_SCOPE_HOJE
    assert tab.agenda_table.rowCount() == 2

    tab.agenda_scope_buttons[tcra_tab_module.AGENDA_SCOPE_7D].click()
    get_app().processEvents()
    assert tab.agenda_table.rowCount() == 3

    tab.agenda_scope_buttons[tcra_tab_module.AGENDA_SCOPE_VENCIDOS].click()
    get_app().processEvents()
    assert tab.agenda_table.rowCount() == 1


def test_tcra_tab_can_create_update_and_delete_tcras(tmp_path, monkeypatch):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    warnings = []

    monkeypatch.setattr(tcra_tab_module, "msg_confirm", lambda *args, **kwargs: True)
    monkeypatch.setattr(
        tcra_tab_module.QMessageBox,
        "warning",
        lambda *args, **kwargs: warnings.append(args[2] if len(args) > 2 else ""),
    )

    tab.in_numero_processo.setText("777/2026")
    tab.in_numero_tcra.setText("TCRA-2026-777")
    tab.in_local.setText("Parque Linear Leste")
    tab.in_endereco.setText("Rua das Acacias")
    tab.in_status.setCurrentText("Em acompanhamento")
    tab.in_prazo_final.setText("03/04/2027")
    tab.in_area_m2.setText("1500,5")
    tab.in_numero_mudas.setText("250")
    tab.in_observacoes.setPlainText("Primeira versao do cadastro")
    tab.save_tcra()

    persisted = service.list_tcras()
    assert len(persisted) == 1
    assert persisted[0].local == "Parque Linear Leste"
    assert persisted[0].area_m2 == 1500.5
    assert tab.current_form_uid == persisted[0].uid

    tab.in_observacoes.setPlainText("Cadastro atualizado")
    tab.save_tcra()

    updated = service.list_tcras()
    assert len(updated) == 1
    assert updated[0].observacoes == "Cadastro atualizado"

    tab.delete_tcra()

    assert service.list_tcras() == []
    assert warnings == []


def test_tcra_tab_audits_create_only_after_save(tmp_path, monkeypatch):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    audit_calls = []

    tab.main_window = SimpleNamespace(
        audit_service=SimpleNamespace(append_session_event=lambda **kwargs: audit_calls.append(kwargs)),
        shell_controller=SimpleNamespace(current_session_path=lambda: "session://banco-local"),
    )
    monkeypatch.setattr(tcra_tab_module, "msg_confirm", lambda *args, **kwargs: True)

    class AddDialog:
        def __init__(self, *args, **kwargs):
            pass

        def exec(self):
            return True

        def values(self):
            return {
                "data_evento": "05/04/2026",
                "tipo_evento": "Vistoria",
                "descricao": "Primeira vistoria registrada.",
                "prazo_resultante": "05/05/2026",
                "status_resultante": "Em acompanhamento",
            }

    monkeypatch.setattr(tcra_tab_module, "TcraEventoEditorDialog", AddDialog)
    tab.in_numero_processo.setText("777/2026")
    tab.in_numero_tcra.setText("TCRA-2026-777")
    tab.in_local.setText("Parque Linear Leste")
    tab.add_event()
    get_app().processEvents()

    assert audit_calls == []

    tab.save_tcra()
    get_app().processEvents()

    assert len(audit_calls) == 1
    assert audit_calls[-1]["session_path"] == "session://banco-local"
    assert audit_calls[-1]["action"] == "TCRA_CREATE"
    assert audit_calls[-1]["metadata"]["event_change_action"] == "add"
    assert audit_calls[-1]["metadata"]["event_change_type"] == "Vistoria"
    assert audit_calls[-1]["after"]["eventos"][0]["tipo_evento"] == "Vistoria"


def test_tcra_tab_audits_event_edit_only_after_save(tmp_path, monkeypatch):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all([make_tcra(uid="tcra-1")])
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()
    audit_calls = []
    warnings = []

    tab.main_window = SimpleNamespace(
        audit_service=SimpleNamespace(append_session_event=lambda **kwargs: audit_calls.append(kwargs)),
        shell_controller=SimpleNamespace(current_session_path=lambda: "session://banco-local"),
    )
    monkeypatch.setattr(
        tcra_tab_module.QMessageBox,
        "warning",
        lambda *args, **kwargs: warnings.append(args[2] if len(args) > 2 else ""),
    )

    class EditDialog:
        def __init__(self, *args, **kwargs):
            pass

        def exec(self):
            return True

        def values(self):
            return {
                "data_evento": "06/04/2026",
                "tipo_evento": "Despacho",
                "descricao": "Evento revisado antes do save.",
                "prazo_resultante": "10/05/2026",
                    "status_resultante": "Relatório pendente",
            }

    tab.table.selectRow(0)
    get_app().processEvents()
    tab.btn_open_selected.click()
    get_app().processEvents()
    tab.events_table.selectRow(0)
    get_app().processEvents()

    monkeypatch.setattr(tcra_tab_module, "TcraEventoEditorDialog", EditDialog)
    tab.edit_selected_event()
    get_app().processEvents()

    assert audit_calls == []

    tab.save_tcra()
    get_app().processEvents()

    assert warnings == []
    assert len(audit_calls) == 1
    assert audit_calls[-1]["action"] == "TCRA_EDIT"
    assert audit_calls[-1]["metadata"]["event_change_action"] == "edit"
    assert audit_calls[-1]["metadata"]["event_change_type"] == "Despacho"
    assert audit_calls[-1]["before"]["eventos"][0]["tipo_evento"] == "Relatório"
    assert audit_calls[-1]["after"]["eventos"][0]["tipo_evento"] == "Despacho"


def test_tcra_tab_can_manage_event_rows(tmp_path, monkeypatch):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))

    class AddDialog:
        def __init__(self, *args, **kwargs):
            pass

        def exec(self):
            return True

        def values(self):
            return {
                "data_evento": "05/04/2026",
                "tipo_evento": "Vistoria",
                "descricao": "Equipe realizou vistoria tecnica.",
                "prazo_resultante": "10/05/2026",
                "status_resultante": "Em acompanhamento",
            }

    class EditDialog:
        def __init__(self, *args, **kwargs):
            pass

        def exec(self):
            return True

        def values(self):
            return {
                "data_evento": "06/04/2026",
                "tipo_evento": "Despacho",
                "descricao": "Despacho atualizado apos vistoria.",
                "prazo_resultante": "15/05/2026",
                "status_resultante": "Relatorio pendente",
            }

    monkeypatch.setattr(tcra_tab_module, "TcraEventoEditorDialog", AddDialog)
    tab.add_event()
    get_app().processEvents()

    assert len(tab.form_eventos) == 1
    assert tab.events_table.rowCount() == 1
    assert tab.form_eventos[0].tipo_evento == "Vistoria"

    monkeypatch.setattr(tcra_tab_module, "TcraEventoEditorDialog", EditDialog)
    tab.events_table.selectRow(0)
    get_app().processEvents()
    tab.edit_selected_event()

    assert tab.form_eventos[0].tipo_evento == "Despacho"
    assert tab.form_eventos[0].status_resultante == "Relatório pendente"

    tab.events_table.selectRow(0)
    get_app().processEvents()
    tab.delete_selected_event()

    assert tab.form_eventos == []
    assert tab.events_table.rowCount() == 0
    assert tab.btn_edit_event.isEnabled() is False


def test_tcra_tab_agenda_selects_record_and_events_update_form(tmp_path, monkeypatch):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(uid="tcra-1", prazo_final=date(2026, 3, 1), data_proximo_relatorio=date(2026, 4, 10)),
            make_tcra(uid="tcra-2", prazo_final=date(2026, 8, 1), data_proximo_relatorio=date(2026, 4, 20)),
        ]
    )
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()

    tab.agenda_table.selectRow(0)
    get_app().processEvents()

    assert tab.current_form_uid == ""
    assert tab.btn_agenda_open.isEnabled() is True

    tab.btn_agenda_open.click()
    get_app().processEvents()

    assert tab.current_form_uid == "tcra-1"

    tab.new_tcra()
    tab.in_numero_processo.setText("777/2026")
    tab.in_local.setText("Parque Linear Leste")
    tab.in_periodicidade.setText("6")

    class ReportDialog:
        def __init__(self, *args, **kwargs):
            pass

        def exec(self):
            return True

        def values(self):
            return {
                "preset_key": "relatorio_entregue",
                "data_evento": "05/04/2026",
                "tipo_evento": "Relatorio entregue",
                "descricao": "Relatorio protocolado.",
                "prazo_resultante": "",
                "status_resultante": "Em acompanhamento",
            }

    class DoneDialog:
        def __init__(self, *args, **kwargs):
            pass

        def exec(self):
            return True

        def values(self):
            return {
                "preset_key": "cumprimento",
                "data_evento": "10/04/2026",
                "tipo_evento": "Cumprimento",
                "descricao": "Termo cumprido.",
                "prazo_resultante": "10/04/2026",
                "status_resultante": "Cumprido",
            }

    monkeypatch.setattr(tcra_tab_module, "TcraEventoEditorDialog", ReportDialog)
    tab.add_event()
    get_app().processEvents()

    assert tab.in_data_ultimo_relatorio.text() == "05/04/2026"
    assert tab.in_data_proximo_relatorio.text() == "05/10/2026"
    assert tab.in_status.currentText() == "Em acompanhamento"

    monkeypatch.setattr(tcra_tab_module, "TcraEventoEditorDialog", DoneDialog)
    tab.events_table.selectRow(0)
    get_app().processEvents()
    tab.edit_selected_event()
    get_app().processEvents()

    assert tab.in_status.currentText() == "Cumprido"
    assert tab.in_data_proximo_relatorio.text() == ""
    assert tab.in_prazo_final.text() == "10/04/2026"


def test_tcra_tab_agenda_quick_actions_update_records_without_opening_editor(tmp_path, monkeypatch):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(
                uid="tcra-1",
                prazo_final=date(2026, 3, 1),
                data_proximo_relatorio=date(2026, 4, 10),
                responsavel_execucao="",
                eventos=[],
            )
        ]
    )

    class InboxEventDialog:
        def __init__(self, *args, **kwargs):
            self.kwargs = kwargs

        def exec(self):
            return True

        def values(self):
            return {
                "preset_key": self.kwargs.get("preset_key", ""),
                "data_evento": "05/04/2026",
                "tipo_evento": "Despacho",
                "descricao": "Cobrança feita pela Inbox.",
                "prazo_resultante": "20/04/2026",
                "status_resultante": "Em acompanhamento",
                "protocolo": "SEI-123",
                "documento_ref": "docs/relatorio.pdf",
            }

    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()
    monkeypatch.setattr(tcra_tab_module, "TcraEventoEditorDialog", InboxEventDialog)

    tab.agenda_table.selectRow(0)
    get_app().processEvents()
    tab.btn_agenda_quick_event.click()
    get_app().processEvents()

    updated = service.get_tcra("tcra-1")
    assert tab.current_form_uid == ""
    assert updated.eventos[0].protocolo == "SEI-123"
    assert updated.eventos[0].documento_ref == "docs/relatorio.pdf"
    assert updated.prazo_final == date(2026, 4, 20)

    monkeypatch.setattr(
        tcra_tab_module.QInputDialog,
        "getText",
        lambda *args, **kwargs: ("Equipe TCRA", True),
    )
    tab.agenda_scope_buttons[tcra_tab_module.AGENDA_SCOPE_7D].click()
    get_app().processEvents()
    tab.agenda_table.selectRow(0)
    get_app().processEvents()
    tab.btn_agenda_assign_responsavel.click()
    get_app().processEvents()

    assert service.get_tcra("tcra-1").responsavel_execucao == "Equipe TCRA"


def test_tcra_tab_blocks_inconsistent_save(tmp_path, monkeypatch):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    warnings = []
    focused_issues = []

    monkeypatch.setattr(
        tcra_tab_module.QMessageBox,
        "warning",
        lambda *args, **kwargs: warnings.append(args[2] if len(args) > 2 else ""),
    )
    monkeypatch.setattr(tab, "_focus_issue_in_form", lambda issue: focused_issues.append(issue))

    tab.in_numero_processo.setText("888/2026")
    tab.in_local.setText("Area Norte")
    tab.in_status.setCurrentText("Cumprido")
    tab.in_data_ultimo_relatorio.setText("10/04/2026")
    tab.in_data_proximo_relatorio.setText("01/04/2026")
    tab.save_tcra()

    assert service.list_tcras() == []
    assert warnings
    assert "Revise o cadastro do TCRA antes de salvar" in warnings[-1]
    assert focused_issues == ["Próximo relatório não pode ser anterior ao último relatório."]


def test_tcra_tab_live_preview_updates_fix_guidance(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))

    tab.in_numero_processo.setText("888/2026")
    tab.in_local.setText("Area Norte")
    tab.in_status.setCurrentText("Cumprido")
    tab.in_data_ultimo_relatorio.setText("10/04/2026")
    tab.in_data_proximo_relatorio.setText("01/04/2026")
    get_app().processEvents()

    assert "sequência cronológica" in tab.lbl_fix_guidance.text().lower()


def test_tcra_tab_quality_selection_routes_to_editor_and_focuses_issue(tmp_path, monkeypatch):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(
                uid="tcra-2",
                numero_tcra="",
                numero_processo="2360/2021",
                local="Varjao",
                bairro="Varjao",
                orgao_acompanhamento="",
                responsavel_execucao="",
                eventos=[],
            )
        ]
    )
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()
    focused = []

    monkeypatch.setattr(tab, "_focus_quality_item", lambda item: focused.append(item.issues[0]))

    tab.quality_table.selectRow(0)
    get_app().processEvents()

    assert tab.workspace_tabs.tabText(tab.workspace_tabs.currentIndex()) == "Cadastro"
    assert focused == ["Sem número TCRA"]


def test_tcra_tab_imports_legacy_workbook_into_local_database(tmp_path, monkeypatch):
    workbook_path = tmp_path / "tcras.xlsx"
    build_legacy_tcra_workbook(workbook_path)
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))

    monkeypatch.setattr(tcra_tab_module, "msg_confirm", lambda *args, **kwargs: True)
    monkeypatch.setattr(
        tcra_tab_module.QFileDialog,
        "getOpenFileName",
        lambda *args, **kwargs: (str(workbook_path), "Planilhas Excel (*.xlsx *.xlsm)"),
    )
    class PreviewDialog:
        def __init__(self, _parent, analysis):
            self.analysis = analysis
        def exec(self):
            return True

    monkeypatch.setattr(tcra_tab_module, "TcraImportPreviewDialog", PreviewDialog)

    tab.import_legacy_workbook()
    get_app().processEvents()
    imported = service.list_tcras()

    assert len(imported) == 2
    assert tab.table.rowCount() == 2
    assert tab.card_total.lbl_value.text() == "2"
    assert any(record.numero_processo == "2360/2021" for record in imported)
    assert "Merge importado: 2" in tab.lbl_import_status.text()
    assert tab.lbl_import_status.isHidden() is False


def test_tcra_tab_audits_import_summary(tmp_path, monkeypatch):
    workbook_path = tmp_path / "tcras.xlsx"
    build_legacy_tcra_workbook(workbook_path)
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    audit_calls = []

    tab.main_window = SimpleNamespace(
        audit_service=SimpleNamespace(append_session_event=lambda **kwargs: audit_calls.append(kwargs)),
        shell_controller=SimpleNamespace(current_session_path=lambda: "session://banco-local"),
    )

    monkeypatch.setattr(tcra_tab_module, "msg_confirm", lambda *args, **kwargs: True)
    monkeypatch.setattr(
        tcra_tab_module.QFileDialog,
        "getOpenFileName",
        lambda *args, **kwargs: (str(workbook_path), "Planilhas Excel (*.xlsx *.xlsm)"),
    )

    class PreviewDialog:
        def __init__(self, _parent, analysis):
            self.analysis = analysis

        def exec(self):
            return True

    monkeypatch.setattr(tcra_tab_module, "TcraImportPreviewDialog", PreviewDialog)

    tab.import_legacy_workbook()
    get_app().processEvents()

    assert audit_calls
    assert audit_calls[-1]["action"] == "TCRA_IMPORT"
    assert audit_calls[-1]["metadata"]["importable_count"] == 2
    assert audit_calls[-1]["metadata"]["mode"] == "merge"
    assert "issue_codes" in audit_calls[-1]["metadata"]


def test_tcra_tab_blocks_duplicate_save(tmp_path, monkeypatch):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all([make_tcra(uid="tcra-1", numero_tcra="TCRA-2026-001", numero_processo="123/2026", local="Area Sul")])
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    warnings = []

    monkeypatch.setattr(
        tcra_tab_module.QMessageBox,
        "warning",
        lambda *args, **kwargs: warnings.append(args[2] if len(args) > 2 else ""),
    )

    tab.new_tcra()
    tab.in_numero_processo.setText("999/2026")
    tab.in_numero_tcra.setText("TCRA-2026-001")
    tab.in_local.setText("Novo local")
    tab.save_tcra()

    assert len(service.list_tcras()) == 1
    assert warnings
    assert "Já existe um TCRA parecido" in warnings[-1]


def test_tcra_tab_quick_filters_update_counts_and_alert_style(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(uid="tcra-1", numero_tcra="", prazo_final=date(2026, 3, 1), data_proximo_relatorio=date(2026, 3, 10)),
            make_tcra(
                uid="tcra-2",
                numero_tcra="TCRA-2026-002",
                prazo_final=date(2026, 8, 1),
                data_proximo_relatorio=date(2026, 4, 20),
                responsavel_execucao="Equipe local",
            ),
            make_tcra(
                uid="tcra-3",
                numero_tcra="TCRA-2026-003",
                status="Cumprido",
                prazo_final=date(2025, 1, 1),
                data_proximo_relatorio=None,
                responsavel_execucao="Equipe local",
            ),
        ]
    )

    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()

    assert "Alertas (1)" in tab.quick_filter_buttons["alertas"].text()
    assert "Próx. 30d (1)" in tab.quick_filter_buttons["proximos"].text()
    assert "Sem número (1)" in tab.quick_filter_buttons["sem_numero"].text()

    tab.quick_filter_buttons["alertas"].click()
    get_app().processEvents()

    assert tab.table.rowCount() == 1
    assert tab.table.item(0, 0).background().color().name().lower() == "#ffffff"
    assert tab.table.item(0, 3).background().color().name().lower() == "#f8cdd3"
    assert tab.table.item(0, 3).foreground().color().name().lower() == "#881337"


def test_tcra_tab_neutral_cells_keep_visible_text_when_theme_changes(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(
                uid="tcra-1",
                numero_tcra="TCRA-2026-003",
                status="Em acompanhamento",
                prazo_final=date(2027, 4, 1),
                data_proximo_relatorio=date(2027, 5, 1),
                responsavel_execucao="Equipe local",
            )
        ]
    )

    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()
    tab.main_window = SimpleNamespace(is_dark_mode=True)
    tab._repaint_table_styles()

    assert tab.table.item(0, 0).background().color().name().lower() == "#0f172a"
    assert tab.table.item(0, 0).foreground().color().name().lower() == "#e5e7eb"

    tab.main_window = SimpleNamespace(is_dark_mode=False)
    tab._repaint_table_styles()

    assert tab.table.item(0, 0).background().color().name().lower() == "#ffffff"
    assert tab.table.item(0, 0).foreground().color().name().lower() == "#111827"
    assert tab.table.item(0, 3).background().color().name().lower() == "#bfdbfe"
    assert tab.table.item(0, 3).foreground().color().name().lower() == "#1e40af"


def test_tcra_tab_light_theme_keeps_status_badges_for_operational_states(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(
                uid="tcra-1",
                status="Em acompanhamento",
                prazo_final=date(2027, 4, 1),
                data_proximo_relatorio=date(2027, 5, 1),
                responsavel_execucao="Equipe local",
            ),
            make_tcra(
                uid="tcra-2",
                status="Cumprido",
                prazo_final=date(2024, 1, 1),
                data_proximo_relatorio=None,
                responsavel_execucao="Equipe local",
            ),
        ]
    )

    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()
    tab.main_window = SimpleNamespace(is_dark_mode=False)
    tab._repaint_table_styles()

    assert tab.table.item(0, 3).background().color().name().lower() == "#bfdbfe"
    assert tab.table.item(0, 3).foreground().color().name().lower() == "#1e40af"
    assert tab.table.item(1, 3).background().color().name().lower() == "#bbf7d0"
    assert tab.table.item(1, 3).foreground().color().name().lower() == "#166534"


def test_tcra_tab_quality_queue_selects_record(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(
                uid="tcra-1",
                numero_tcra="",
                responsavel_execucao="",
                orgao_acompanhamento="",
                local="Area Norte",
            ),
            make_tcra(
                uid="tcra-2",
                numero_tcra="TCRA-2026-002",
                responsavel_execucao="Equipe local",
                orgao_acompanhamento="CETESB",
                local="Parque Linear",
            ),
        ]
    )

    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()

    assert tab.quality_table.rowCount() == 1
    assert "Qualidade cadastral:" in tab.lbl_quality_summary.text()

    tab.quality_table.selectRow(0)
    get_app().processEvents()

    assert tab.current_form_uid == "tcra-1"
    assert tab.in_local.text() == "Area Norte"
    assert tab.workspace_tabs.tabText(tab.workspace_tabs.currentIndex()) == "Cadastro"
    assert tab.form_panel_body.isHidden() is False


def test_tcra_tab_open_selected_button_switches_to_cadastro_and_keeps_form_visible(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all([make_tcra(uid="tcra-1")])
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()

    tab.table.selectRow(0)
    get_app().processEvents()

    assert tab.btn_open_selected.isEnabled() is True
    assert tab.btn_record_details.isEnabled() is True
    assert tab.selection_actions_frame.isHidden() is False
    assert tab.current_form_uid == ""
    assert tab.operational_dialog.isVisible() is False
    assert tab.record_details.toPlainText()
    assert tab.workspace_tabs.tabText(tab.workspace_tabs.currentIndex()) == "Lista"

    tab.btn_open_selected.click()
    get_app().processEvents()

    assert tab.workspace_tabs.tabText(tab.workspace_tabs.currentIndex()) == "Cadastro"
    assert tab.editor_tabs.currentIndex() == 0
    assert tab.form_panel_body.isHidden() is False
    assert tab.btn_toggle_form_panel.isChecked() is True
    assert tab.editor_operational_panel.isHidden() is True


def test_tcra_tab_record_details_opens_in_dialog_without_side_panel(tmp_path, monkeypatch):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all([make_tcra(uid="tcra-1")])
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()
    opened = []

    class DetailsDialog:
        edit_requested = False

        def __init__(self, parent, **kwargs):
            opened.append((kwargs["record"].uid, kwargs["today"]))

        def exec(self):
            opened.append(("exec", None))
            return True

    monkeypatch.setattr(tcra_tab_module, "TcraRecordDetailsDialog", DetailsDialog)

    tab.table.selectRow(0)
    get_app().processEvents()
    assert tab.operational_dialog.isVisible() is False

    tab.btn_record_details.click()
    get_app().processEvents()

    assert opened == [("tcra-1", date(2026, 4, 3)), ("exec", None)]
    assert tab.workspace_tabs.tabText(tab.workspace_tabs.currentIndex()) == "Lista"
    assert tab.operational_dialog.isVisible() is False


def test_tcra_tab_quick_event_buttons_apply_presets(tmp_path, monkeypatch):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all([make_tcra(uid="tcra-1")])
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()
    requested_presets = []

    monkeypatch.setattr(
        tab,
        "_open_current_form_record_details",
        lambda *, event_preset="": requested_presets.append(event_preset),
    )

    tab.table.selectRow(0)
    get_app().processEvents()
    tab.btn_quick_report.click()
    get_app().processEvents()
    tab.btn_quick_done.click()
    get_app().processEvents()

    assert requested_presets == ["relatorio_entregue", "cumprimento"]


def test_tcra_tab_exports_excel_and_pdf_reports(tmp_path, monkeypatch):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all([make_tcra(uid="tcra-1"), make_tcra(uid="tcra-2", status="Cumprido", data_proximo_relatorio=None)])
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()
    info_messages = []
    save_paths = iter(
        [
            (str(tmp_path / "tcra-export.xlsx"), "Planilha (*.xlsx)"),
            (str(tmp_path / "tcra-export.pdf"), "PDF (*.pdf)"),
        ]
    )

    monkeypatch.setattr(tcra_tab_module.QFileDialog, "getSaveFileName", lambda *args, **kwargs: next(save_paths))
    monkeypatch.setattr(
        tcra_tab_module.QMessageBox,
        "information",
        lambda *args, **kwargs: info_messages.append(args[2] if len(args) > 2 else ""),
    )

    tab.export_excel_report()
    tab.export_pdf_report()

    workbook = load_workbook(tmp_path / "tcra-export.xlsx")
    assert workbook.sheetnames == ["Resumo", "TCRAs"]
    assert (tmp_path / "tcra-export.pdf").exists() is True
    assert (tmp_path / "tcra-export.pdf").stat().st_size > 0
    assert len(info_messages) == 2


def test_tcra_tab_restores_and_persists_filter_state(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(uid="tcra-1", bairro="Centro", orgao_acompanhamento="MPSP", numero_tcra="", responsavel_execucao=""),
            make_tcra(uid="tcra-2", bairro="Varjao", orgao_acompanhamento="CETESB", responsavel_execucao=""),
        ]
    )
    saved_states = []
    settings_controller = SimpleNamespace(
        tcra_filter_state=lambda: {
            "search_text": "varjao",
            "status": "Todos",
            "selected_orgaos": ["CETESB"],
            "orgaos_all_selected": False,
            "selected_bairros": ["Varjao"],
            "bairros_all_selected": False,
            "year": "Todos",
            "only_mpsp": False,
            "only_relatorio_pendente": False,
            "only_prazo_vencido": False,
            "quick_filter_mode": "sem_responsavel",
        },
        set_tcra_filter_state=lambda state: saved_states.append(dict(state)),
        preferred_export_dir=lambda: "",
        save_last_export_dir=lambda _path: None,
    )

    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    tab.main_window = SimpleNamespace(settings_controller=settings_controller)
    tab._pending_filter_restore = settings_controller.tcra_filter_state()
    tab.refresh_data()
    get_app().processEvents()

    assert tab.search_input.text() == "varjao"
    assert tab.quick_filter_mode == "sem_responsavel"
    assert tab.table.rowCount() == 1

    tab.quick_filter_buttons["all"].click()
    tab.chk_only_mpsp.setChecked(True)
    get_app().processEvents()

    assert saved_states
    assert saved_states[-1]["quick_filter_mode"] == "all"
    assert saved_states[-1]["only_mpsp"] is True
    assert "selected_responsaveis" in saved_states[-1]
    assert "responsaveis_all_selected" in saved_states[-1]


def test_tcra_tab_bulk_action_updates_selected_records(tmp_path, monkeypatch):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    service.replace_all(
        [
            make_tcra(uid="tcra-1", prazo_final=date(2026, 3, 1), data_proximo_relatorio=date(2026, 4, 1), responsavel_execucao=""),
            make_tcra(uid="tcra-2", prazo_final=date(2026, 3, 2), data_proximo_relatorio=date(2026, 4, 2), responsavel_execucao=""),
        ]
    )

    class BulkDialog:
        def __init__(self, *args, **kwargs):
            pass

        def exec(self):
            return True

        def values(self):
            return {
                "action": "responsavel",
                "status": "",
                "text_value": "Equipe TCRA",
                "date_value": "",
                "event_preset": "",
                "event_date": "",
                "event_deadline": "",
            }

    monkeypatch.setattr(tcra_tab_module, "TcraBulkActionDialog", BulkDialog)

    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    get_app().processEvents()
    tab._select_alert_rows()
    get_app().processEvents()

    tab.apply_bulk_action()

    assert service.get_tcra("tcra-1").responsavel_execucao == "Equipe TCRA"
    assert service.get_tcra("tcra-2").responsavel_execucao == "Equipe TCRA"


def test_tcra_tab_restores_new_form_draft_from_settings(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    settings_controller = SimpleNamespace(
        tcra_form_draft=lambda: {
            "uid": "",
            "numero_processo": "999/2026",
            "numero_tcra": "",
            "local": "Area em rascunho",
            "endereco": "",
            "bairro": "",
            "orgao": "",
            "status": "Em acompanhamento",
            "data_assinatura": "",
            "prazo_final": "",
            "periodicidade": "",
            "data_ultimo_relatorio": "",
            "data_proximo_relatorio": "",
            "area_m2": "",
            "numero_mudas": "",
            "responsavel": "",
            "mpsp": False,
            "inquerito": "",
            "servicos": "Servicos em rascunho",
            "observacoes": "",
            "eventos": (),
        },
        set_tcra_form_draft=lambda _state: None,
        clear_tcra_form_draft=lambda: None,
        preferred_export_dir=lambda: "",
        save_last_export_dir=lambda _path: None,
        set_tcra_filter_state=lambda _state: None,
        tcra_filter_state=lambda: {},
    )

    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))
    tab.main_window = SimpleNamespace(settings_controller=settings_controller)
    tab._pending_new_form_draft = settings_controller.tcra_form_draft()

    tab.new_tcra()
    get_app().processEvents()

    assert tab.in_numero_processo.text() == "999/2026"
    assert tab.in_local.text() == "Area em rascunho"
    assert tab.has_pending_form_changes() is True


def test_tcra_tab_global_search_mode_hides_local_search(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))

    tab.set_global_search_mode(True)
    assert tab.search_input.isHidden() is True
    assert tab.lbl_search.isHidden() is True

    tab.set_global_search_mode(False)
    assert tab.search_input.isHidden() is False
    assert tab.lbl_search.isHidden() is False


def test_tcra_tab_safe_fix_clears_open_next_report_for_completed_term(tmp_path):
    service = TcraSqliteService(db_path=tmp_path / "local.db")
    tab = TcraTab(sqlite_service=service, today=date(2026, 4, 3))

    tab.in_numero_processo.setText("999/2026")
    tab.in_status.setCurrentText("Cumprido")
    tab.in_data_proximo_relatorio.setText("10/05/2026")
    get_app().processEvents()

    tab._apply_safe_fix()

    assert tab.in_data_proximo_relatorio.text() == ""
